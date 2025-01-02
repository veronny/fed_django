from django.shortcuts import render

# TABLERO Adolescente Hemopglobina 
from django.db import connection
from django.http import JsonResponse
from base.models import MAESTRO_HIS_ESTABLECIMIENTO, DimPeriodo
from django.db.models.functions import Substr
import logging

# report excel
from django.http.response import HttpResponse
from django.views.generic.base import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import openpyxl
from openpyxl.utils import get_column_letter

from django.db.models.functions import Substr

from datetime import datetime
import locale

from django.db.models import IntegerField  # Importar IntegerField
from django.db.models.functions import Cast, Substr  # Importar Cast y Substr


logger = logging.getLogger(__name__)


# Reporte excel
from datetime import datetime
import getpass  # Para obtener el nombre del usuario
from django.contrib.auth.models import User  # O tu modelo de usuario personalizado
from django.http import HttpResponse
from io import BytesIO
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required

User = get_user_model()

from django.db.models import IntegerField               # Importar IntegerField
from django.db.models.functions import Cast, Substr     # Importar Cast y Substr
from base.models import Actualizacion
# Create your views here.
def obtener_distritos(provincia):
    distritos = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Provincia=provincia).values('Distrito').distinct().order_by('Distrito')
    return list(distritos)

def obtener_avance_s23_suplementacion12(red):
    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT * FROM public.obtener_avance_s23_suplementacion12(%s)",
            [red]
        )
        return cursor.fetchall()

def obtener_ranking_s23_suplementacion12(anio, mes):
    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT * FROM public.obtener_ranking_s23_suplementacion12(%s, %s)",
            [anio, mes]
        )
        result = cursor.fetchall()
        return result

## AVANCE REGIONAL
def obtener_avance_regional_s23_suplementacion12():
    """
    Obtiene el avance regional de gestantes con anemia.
    Retorna una lista de diccionarios con las claves 'num', 'den' y 'cob'.
    """
    try:
        # Asegúrate de que la conexión a la base de datos está establecida
        with connection.cursor() as cursor:
            cursor.execute(
                '''
                SELECT 
                    SUM(numerador_mes_diciembre) AS num,
                    SUM(denominador_mes_diciembre) AS den,
                    ROUND((SUM(numerador_mes_diciembre)::NUMERIC / NULLIF(SUM(denominador_mes_diciembre), 0)) * 100, 2) AS cob
                FROM public."Cobertura_SI_0203_Suplementacion12M"
                '''
            )
            resultados = cursor.fetchall()
            
            # Obtener los nombres de las columnas
            column_names = [desc[0] for desc in cursor.description]
            
            # Convertir cada fila en un diccionario
            datos = [dict(zip(column_names, fila)) for fila in resultados]
        
        return datos
    except Exception as e:
        print(f"Error al obtener el avance regional: {e}")
        return None


## AVANCE REGIONAL MENSUALIZADO
def obtener_avance_regional_mensual_s23_suplementacion12():
    """
    Obtiene el avance regional de gestantes con anemia de manera mensualizada.
    Retorna una lista de diccionarios con las claves 'num', 'den' y 'cob' por meses.
    """
    try:
        # Asegúrate de que la conexión a la base de datos está establecida
        with connection.cursor() as cursor:
            cursor.execute(
                '''
                    SELECT 
                    SUM(numerador_mes_enero) AS NUM_1, 
                    SUM(denominador_mes_enero) AS DEN_1, 
                    ROUND((SUM(numerador_mes_enero)::NUMERIC / NULLIF(SUM(denominador_mes_enero), 0)) * 100, 2) AS COB_1,
                    SUM(numerador_mes_febrero) AS NUM_2, 
                    SUM(denominador_mes_febrero) AS DEN_2, 
                    ROUND((SUM(numerador_mes_febrero)::NUMERIC / NULLIF(SUM(denominador_mes_febrero), 0)) * 100, 2) AS COB_2,
                    SUM(numerador_mes_marzo) AS NUM_3, 
                    SUM(denominador_mes_marzo) AS DEN_3, 
                    ROUND((SUM(numerador_mes_marzo)::NUMERIC / NULLIF(SUM(denominador_mes_marzo), 0)) * 100, 2) AS COB_3,
                    SUM(numerador_mes_abril) AS NUM_4, 
                    SUM(denominador_mes_abril) AS DEN_4, 
                    ROUND((SUM(numerador_mes_abril)::NUMERIC / NULLIF(SUM(denominador_mes_abril), 0)) * 100, 2) AS COB_4,
                    SUM(numerador_mes_mayo) AS NUM_5, 
                    SUM(denominador_mes_mayo) AS DEN_5, 
                    ROUND((SUM(numerador_mes_mayo)::NUMERIC / NULLIF(SUM(denominador_mes_mayo), 0)) * 100, 2) AS COB_5,
                    SUM(numerador_mes_junio) AS NUM_6, 
                    SUM(denominador_mes_junio) AS DEN_6,
                    ROUND((SUM(numerador_mes_junio)::NUMERIC / NULLIF(SUM(denominador_mes_junio), 0)) * 100, 2) AS COB_6,
                    SUM(numerador_mes_julio) AS NUM_7, 
                    SUM(denominador_mes_julio) AS DEN_7,
                    ROUND((SUM(numerador_mes_julio)::NUMERIC / NULLIF(SUM(denominador_mes_julio), 0)) * 100, 2) AS COB_7,
                    SUM(numerador_mes_agosto) AS NUM_8, 
                    SUM(denominador_mes_agosto) AS DEN_8, 
                    ROUND((SUM(numerador_mes_agosto)::NUMERIC / NULLIF(SUM(denominador_mes_agosto), 0)) * 100, 2) AS COB_8,
                    SUM(numerador_mes_setiembre) AS NUM_9, 
                    SUM(denominador_mes_setiembre) AS DEN_9, 
                    ROUND((SUM(numerador_mes_setiembre)::NUMERIC / NULLIF(SUM(denominador_mes_setiembre), 0)) * 100, 2) AS COB_9,
                    SUM(numerador_mes_octubre) AS NUM_10, 
                    SUM(denominador_mes_octubre) AS DEN_10, 
                    ROUND((SUM(numerador_mes_octubre)::NUMERIC / NULLIF(SUM(denominador_mes_octubre), 0)) * 100, 2) AS COB_10,
                    SUM(numerador_mes_noviembre) AS NUM_11, 
                    SUM(denominador_mes_noviembre) AS DEN_11, 
                    ROUND((SUM(numerador_mes_noviembre)::NUMERIC / NULLIF(SUM(denominador_mes_noviembre), 0)) * 100, 2) AS COB_11,
                    SUM(numerador_mes_diciembre) AS NUM_12, 
                    SUM(denominador_mes_diciembre) AS DEN_12, 
                    ROUND((SUM(numerador_mes_diciembre)::NUMERIC / NULLIF(SUM(denominador_mes_diciembre), 0)) * 100, 2) AS COB_12
                    FROM 
                    public."Cobertura_SI_0203_Suplementacion12M";
                '''
            )
            resultados = cursor.fetchall()
            
            # Obtener los nombres de las columnas
            column_names = [desc[0] for desc in cursor.description]
            
            # Convertir cada fila en un diccionario
            datos = [dict(zip(column_names, fila)) for fila in resultados]
        
        return datos
    except Exception as e:
        print(f"Error al obtener el avance regional: {e}")
        return None

def index_s23_suplementacion12(request):
    actualizacion = Actualizacion.objects.all()
    # RANKING 
    anio = request.GET.get('anio')  # Valor predeterminado# Valor predeterminado
    mes_seleccionado = request.GET.get('mes')
    # GRAFICO
    red_seleccionada = request.GET.get('red')
    red = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Disa='JUNIN').values_list('Red', flat=True).distinct().order_by('Red')
    # Si la solicitud es AJAX
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            # Obtener datos de RANKING 
            resultados_ranking_obtener_s23_suplementacion12 = obtener_ranking_s23_suplementacion12(anio,mes_seleccionado)
            # Obtener datos de AVANCE GRAFICO MESES
            resultados_avance_obtener_s23_suplementacion12 = obtener_avance_s23_suplementacion12(red_seleccionada)
            # Obtener datos de AVANCE GRAFICO ANUAL
            resultados_avance_regional_s23_suplementacion12 = obtener_avance_regional_s23_suplementacion12()
            # Obtener datos de AVANCE GRAFICO ANUAL
            resultados_avance_regional_mensual_s23_suplementacion12 = obtener_avance_regional_mensual_s23_suplementacion12()
            # Procesar los resultados
            if any(len(row) < 4 for row in resultados_ranking_obtener_s23_suplementacion12):
                raise ValueError("Algunas filas del ranking no tienen suficientes elementos")
            
            data = {               
                #ranking
                'red': [],
                'num_r': [],
                'den_r': [],
                'avance_r': [],
                
                #avance meses
                'mes': [],
                'num': [],
                'den': [],
                'avance': [],
                
            #avance regional
                'num_region': [],
                'den_region': [],
                'avance_region': [],
                
                #avance regional mensual
                'num_1': [],
                'den_1': [],
                'cob_1': [],
                'num_2': [],
                'den_2': [],
                'cob_2': [],
                'num_3': [],
                'den_3': [],
                'cob_3': [],
                'num_4': [],
                'den_4': [],
                'cob_4': [],
                'num_5': [],
                'den_5': [],
                'cob_5': [],
                'num_6': [],
                'den_6': [],
                'cob_6': [],
                'num_7': [],
                'den_7': [],
                'cob_7': [],
                'num_8': [],
                'den_8': [],
                'cob_8': [],                
                'num_9': [],
                'den_9': [],
                'cob_9': [],
                'num_10': [],
                'den_10': [],
                'cob_10': [],
                'num_11': [],
                'den_11': [],
                'cob_11': [],
                'num_12': [],
                'den_12': [],
                'cob_12': [],
            }     
            # AVANCE GRAFICO REGIONAL
            for index, row in enumerate(resultados_avance_regional_s23_suplementacion12):
                try:
                    # Verifica que el diccionario tenga las claves necesarias
                    required_keys = {'num', 'den', 'cob'}
                    if not required_keys.issubset(row.keys()):
                        raise ValueError(f"La fila {index} no tiene las claves necesarias: {row}")

                    num_region_value = float(row.get('num', 0.0))
                    den_region_value = float(row.get('den', 0.0))
                    avance_region_value = float(row.get('cob', 0.0))

                    data['num_region'].append(num_region_value)
                    data['den_region'].append(den_region_value)
                    data['avance_region'].append(avance_region_value)

                except Exception as e:
                    logger.error(f"Error procesando la fila {index}: {str(e)}")
                    
            # AVANCE GRAFICO MENSUAL
            for index, row in enumerate(resultados_avance_regional_mensual_s23_suplementacion12):
                try:
                    # Verifica que el diccionario tenga las claves necesarias
                    required_keys = {'num_1','den_1','cob_1','num_2','den_2','cob_2','num_3','den_3','cob_3','num_4','den_4','cob_4','num_5','den_5','cob_5','num_6','den_6','cob_6','num_7','den_7','cob_7','num_8','den_8','cob_8','num_9','den_9','cob_9','num_10','den_10','cob_10','num_11','den_11','cob_11','num_12','den_12','cob_12'}
                    
                    if not required_keys.issubset(row.keys()):
                        raise ValueError(f"La fila {index} no tiene las claves necesarias: {row}")

                    num_1_value = float(row.get('num_1', 0.0))
                    den_1_value = float(row.get('den_1', 0.0))
                    cob_1_value = float(row.get('cob_1', 0.0))
                    num_2_value = float(row.get('num_2', 0.0))
                    den_2_value = float(row.get('den_2', 0.0))
                    cob_2_value = float(row.get('cob_2', 0.0))
                    num_3_value = float(row.get('num_3', 0.0))
                    den_3_value = float(row.get('den_3', 0.0))
                    cob_3_value = float(row.get('cob_3', 0.0))
                    num_4_value = float(row.get('num_4', 0.0))
                    den_4_value = float(row.get('den_4', 0.0))
                    cob_4_value = float(row.get('cob_4', 0.0))
                    num_5_value = float(row.get('num_5', 0.0))
                    den_5_value = float(row.get('den_5', 0.0))
                    cob_5_value = float(row.get('cob_5', 0.0))
                    num_6_value = float(row.get('num_6', 0.0))
                    den_6_value = float(row.get('den_6', 0.0))
                    cob_6_value = float(row.get('cob_6', 0.0))
                    num_7_value = float(row.get('num_7', 0.0))
                    den_7_value = float(row.get('den_7', 0.0))
                    cob_7_value = float(row.get('cob_7', 0.0))
                    num_8_value = float(row.get('num_8', 0.0))
                    den_8_value = float(row.get('den_8', 0.0))
                    cob_8_value = float(row.get('cob_8', 0.0))
                    num_9_value = float(row.get('num_9', 0.0))
                    den_9_value = float(row.get('den_9', 0.0))
                    cob_9_value = float(row.get('cob_9', 0.0))
                    num_10_value = float(row.get('num_10', 0.0))
                    den_10_value = float(row.get('den_10', 0.0))
                    cob_10_value = float(row.get('cob_10', 0.0))
                    num_11_value = float(row.get('num_11', 0.0))
                    den_11_value = float(row.get('den_11', 0.0))
                    cob_11_value = float(row.get('cob_11', 0.0))
                    num_12_value = float(row.get('num_12', 0.0))
                    den_12_value = float(row.get('den_12', 0.0))
                    cob_12_value = float(row.get('cob_12', 0.0))
                    
                    data['num_1'].append(num_1_value)
                    data['den_1'].append(den_1_value)
                    data['cob_1'].append(cob_1_value)
                    data['num_2'].append(num_2_value)
                    data['den_2'].append(den_2_value)
                    data['cob_2'].append(cob_2_value)
                    data['num_3'].append(num_3_value)
                    data['den_3'].append(den_3_value)
                    data['cob_3'].append(cob_3_value)
                    data['num_4'].append(num_4_value)
                    data['den_4'].append(den_4_value)
                    data['cob_4'].append(cob_4_value)
                    data['num_5'].append(num_5_value)
                    data['den_5'].append(den_5_value)
                    data['cob_5'].append(cob_5_value)
                    data['num_6'].append(num_6_value)
                    data['den_6'].append(den_6_value)
                    data['cob_6'].append(cob_6_value)
                    data['num_7'].append(num_7_value)
                    data['den_7'].append(den_7_value)
                    data['cob_7'].append(cob_7_value)
                    data['num_8'].append(num_8_value)
                    data['den_8'].append(den_8_value)
                    data['cob_8'].append(cob_8_value)
                    data['num_9'].append(num_9_value)
                    data['den_9'].append(den_9_value)
                    data['cob_9'].append(cob_9_value)
                    data['num_10'].append(num_10_value)
                    data['den_10'].append(den_10_value)
                    data['cob_10'].append(cob_10_value)
                    data['num_11'].append(num_11_value)
                    data['den_11'].append(den_11_value)
                    data['cob_11'].append(cob_11_value)
                    data['num_12'].append(num_12_value)
                    data['den_12'].append(den_12_value)
                    data['cob_12'].append(cob_12_value)

                except Exception as e:
                    logger.error(f"Error procesando la fila {index}: {str(e)}")    
            #RANKING
            for index, row in enumerate(resultados_ranking_obtener_s23_suplementacion12):
                try:
                    # Verifica que la tupla tenga exactamente 4 elementos
                    if len(row) != 4:
                        raise ValueError(f"La fila {index} no tiene 4 elementos: {row}")

                    red_value = row[0] if row[0] is not None else ''
                    num_r_value = float(row[1]) if row[1] is not None else 0.0
                    den_r_value = float(row[2]) if row[2] is not None else 0.0
                    avance_r_value = float(row[3]) if row[3] is not None else 0.0

                    data['red'].append(red_value)
                    data['num_r'].append(num_r_value)
                    data['den_r'].append(den_r_value)
                    data['avance_r'].append(avance_r_value)

                except Exception as e:
                    logger.error(f"Error procesando la fila {index}: {str(e)}")
            
            #AVANCE GRAFICO MESES
            for index, row in enumerate(resultados_avance_obtener_s23_suplementacion12):
                try:
                    # Verifica que la tupla tenga exactamente 4 elementos
                    if len(row) != 5:
                        raise ValueError(f"La fila {index} no tiene 5 elementos: {row}")

                    mes_value = row[1] if row[1] is not None else ''
                    num_value = float(row[2]) if row[2] is not None else 0.0
                    den_value = float(row[3]) if row[3] is not None else 0.0
                    avance_value = float(row[4]) if row[4] is not None else 0.0

                    data['mes'].append(mes_value)
                    data['num'].append(num_value)
                    data['den'].append(den_value)
                    data['avance'].append(avance_value)

                except Exception as e:
                    logger.error(f"Error procesando la fila {index}: {str(e)}")
            
            return JsonResponse(data)

        except Exception as e:
            logger.error(f"Error al obtener datos: {str(e)}")

    # Si no es una solicitud AJAX, renderiza la página principal
    return render(request, 's23_suplementacion12/index_s23_suplementacion12.html', {
        'red': red,
        'mes_seleccionado': mes_seleccionado,
        'actualizacion': actualizacion
    })

## SEGUIMIENTO
def get_redes_s23_suplementacion12(request,redes_id):
    redes = (
            MAESTRO_HIS_ESTABLECIMIENTO
            .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL',Departamento='JUNIN')
            .annotate(codigo_red_filtrado=Substr('Codigo_Red', 1, 4))
            .values('Red','codigo_red_filtrado')
            .distinct()
            .order_by('Red')
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'redes': redes,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
    }
    
    return render(request, 's23_suplementacion12/redes.html', context)

## SEGUIMIENTO POR MICRO-REDES
def get_microredes_s23_suplementacion12(request, microredes_id):
    redes = (
            MAESTRO_HIS_ESTABLECIMIENTO
            .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL',Departamento='JUNIN')
            .annotate(codigo_red_filtrado=Substr('Codigo_Red', 1, 4))
            .values('Red','codigo_red_filtrado')
            .distinct()
            .order_by('Red')
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'redes': redes,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
    }
    
    return render(request, 's23_suplementacion12/microredes.html', context)

def p_microredes_s23_suplementacion12(request):
    redes_param = request.GET.get('red')
    microredes = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Codigo_Red=redes_param, Descripcion_Sector='GOBIERNO REGIONAL', Disa='JUNIN').values('Codigo_MicroRed','MicroRed').distinct()
    context = {
        'redes_param': redes_param,
        'microredes': microredes
    }
    return render(request, 's23_suplementacion12/partials/p_microredes.html', context)

## REPORTE POR ESTABLECIMIENTO
def get_establecimientos_s23_suplementacion12(request,establecimiento_id):
    redes = (
                MAESTRO_HIS_ESTABLECIMIENTO
                .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL',Disa='JUNIN')
                .annotate(codigo_red_filtrado=Substr('Codigo_Red', 1, 4))
                .values('Red','codigo_red_filtrado')
                .distinct()
                .order_by('Red')
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'redes': redes,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
    }
    return render(request,'s23_suplementacion12/establecimientos.html', context)

def p_microredes_establec_s23_suplementacion12(request):
    redes_param = request.GET.get('red') 
    microredes = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Codigo_Red=redes_param, Descripcion_Sector='GOBIERNO REGIONAL',Disa='JUNIN').values('Codigo_MicroRed','MicroRed').distinct()
    context = {
        'microredes': microredes,
        'is_htmx': True
    }
    return render(request, 's23_suplementacion12/partials/p_microredes_establec.html', context)

def p_establecimientos_s23_suplementacion12(request):
    microredes = request.GET.get('p_microredes_establec')    
    codigo_red = request.GET.get('red')
    establec = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Codigo_MicroRed=microredes,Codigo_Red=codigo_red,Descripcion_Sector='GOBIERNO REGIONAL',Disa='JUNIN').values('Codigo_Unico','Nombre_Establecimiento').distinct()

    context= {
        'establec': establec
    }
    return render(request, 's23_suplementacion12/partials/p_establecimientos.html', context)

## QUERY
def obtener_seguimiento_redes_s23_suplementacion12(p_anio,p_red,p_microred,p_establec,p_inicio,p_fin,p_cumple):
    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT * FROM public.fn_seguimiento_s23_suplementacion12(%s,%s,%s,%s,%s,%s,%s)",
            [p_anio, p_red, p_microred, p_establec, p_inicio, p_fin, p_cumple]
        )
        return cursor.fetchall()


## REPORTESA
class RptS23Suplementacion12Red(TemplateView):
    def get(self, request, *args, **kwargs):
        # Variables ingresadas
        p_anio = request.GET.get('anio')
        p_red = request.GET.get('red','')
        p_microred = ''
        p_establec = ''
        p_inicio = int(request.GET.get('fecha_inicio'))
        p_fin = int(request.GET.get('fecha_fin'))
        p_cumple = request.GET.get('cumple', '')  
        # Creación de la consulta
        resultado_seguimiento = obtener_seguimiento_redes_s23_suplementacion12(p_anio,p_red,p_microred,p_establec,p_inicio,p_fin,p_cumple)
        
        wb = Workbook()
        
        consultas = [
                ('Seguimiento', resultado_seguimiento)
        ]
        
        for index, (sheet_name, results) in enumerate(consultas):
            if index == 0:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
        
            fill_worksheet_s23_suplementacion12(ws, results)
        ##########################################################################          
        # Establecer el nombre del archivo
        nombre_archivo = "rpt_s23_suplementacion_12.xlsx"
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)

        return response

class RptS23Suplementacion12MicroRed(TemplateView):
    def get(self, request, *args, **kwargs):
        # Variables ingresadas
        p_anio = request.GET.get('anio')
        p_red = request.GET.get('red','')
        p_microred = request.GET.get('p_microredes','')
        p_establec = ''
        p_inicio = int(request.GET.get('fecha_inicio'))
        p_fin = int(request.GET.get('fecha_fin'))
        p_cumple = request.GET.get('cumple', '')     
        # Creación de la consulta
        resultado_seguimiento = obtener_seguimiento_redes_s23_suplementacion12(p_anio,p_red,p_microred,p_establec,p_inicio,p_fin,p_cumple)
                
        wb = Workbook()
        
        consultas = [
                ('Seguimiento', resultado_seguimiento)
        ]
        
        for index, (sheet_name, results) in enumerate(consultas):
            if index == 0:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
        
            fill_worksheet_s23_suplementacion12(ws, results)
        
        ##########################################################################          
        # Establecer el nombre del archivo
        nombre_archivo = "rpt_s23_suplementacion12_red.xlsx"
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)

        return response

class RptS23Suplementacion12Establec(TemplateView):
    def get(self, request, *args, **kwargs):
        # Variables ingresadas
        p_anio = request.GET.get('anio')
        p_red = request.GET.get('red','')
        p_microred = request.GET.get('p_microredes','')
        p_establec = request.GET.get('p_establecimiento','')
        p_inicio = int(request.GET.get('fecha_inicio'))
        p_fin = int(request.GET.get('fecha_fin'))
        p_cumple = request.GET.get('cumple', '')     
        # Creación de la consulta
        resultado_seguimiento = obtener_seguimiento_redes_s23_suplementacion12(p_anio,p_red,p_microred,p_establec,p_inicio,p_fin,p_cumple)
                
        wb = Workbook()
        
        consultas = [
                ('Seguimiento', resultado_seguimiento)
        ]
        
        for index, (sheet_name, results) in enumerate(consultas):
            if index == 0:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
        
            fill_worksheet_s23_suplementacion12(ws, results)
        
        ##########################################################################          
        # Establecer el nombre del archivo
        nombre_archivo = "rpt_s23_suplementacion12_red.xlsx"
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)

        return response


def fill_worksheet_s23_suplementacion12(ws, results): 
    # cambia el alto de la columna   
    ws.row_dimensions[1].height = 14
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 12
    ws.row_dimensions[4].height = 25
    ws.row_dimensions[5].height = 18
    ws.row_dimensions[6].height = 22
    ws.row_dimensions[7].height = 60
    ws.row_dimensions[8].height = 40
    # cambia el ancho de la columna
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 3
    ws.column_dimensions['C'].width = 9
    ws.column_dimensions['D'].width = 9
    ws.column_dimensions['E'].width = 5
    ws.column_dimensions['F'].width = 9
    ws.column_dimensions['G'].width = 9
    ws.column_dimensions['H'].width = 5
    ws.column_dimensions['I'].width = 9
    ws.column_dimensions['J'].width = 5
    ws.column_dimensions['K'].width = 9
    ws.column_dimensions['L'].width = 5
    ws.column_dimensions['M'].width = 9
    ws.column_dimensions['N'].width = 5
    ws.column_dimensions['O'].width = 9
    ws.column_dimensions['P'].width = 5
    ws.column_dimensions['Q'].width = 11
    ws.column_dimensions['R'].width = 9
    ws.column_dimensions['S'].width = 5
    ws.column_dimensions['T'].width = 9
    ws.column_dimensions['U'].width = 5
    ws.column_dimensions['V'].width = 9
    ws.column_dimensions['W'].width = 5
    ws.column_dimensions['X'].width = 9
    ws.column_dimensions['Y'].width = 5
    ws.column_dimensions['Z'].width = 9
    ws.column_dimensions['AA'].width = 5
    ws.column_dimensions['AB'].width = 9
    ws.column_dimensions['AC'].width = 5
    ws.column_dimensions['AD'].width = 9
    ws.column_dimensions['AE'].width = 5
    ws.column_dimensions['AF'].width = 11
    ws.column_dimensions['AG'].width = 9
    ws.column_dimensions['AH'].width = 5
    ws.column_dimensions['AI'].width = 9
    ws.column_dimensions['AJ'].width = 5
    ws.column_dimensions['AK'].width = 11
    ws.column_dimensions['AL'].width = 11
    ws.column_dimensions['AM'].width = 11
    ws.column_dimensions['AN'].width = 11  
    ws.column_dimensions['AO'].width = 11
    ws.column_dimensions['AP'].width = 11    
    ws.column_dimensions['AQ'].width = 9
    ws.column_dimensions['AR'].width = 16
    ws.column_dimensions['AS'].width = 16
    ws.column_dimensions['AT'].width = 20
    ws.column_dimensions['AU'].width = 20
    ws.column_dimensions['AV'].width = 6
    ws.column_dimensions['AW'].width = 25

    
    # linea de division
    ws.freeze_panes = 'K10'
    # Configuración del fondo y el borde
    # Definir el color usando formato aRGB (opacidad completa 'FF' + color RGB)
    fill = PatternFill(start_color='FF60D7E0', end_color='FF60D7E0', fill_type='solid')
    # Definir el color anaranjado usando formato aRGB
    orange_fill = PatternFill(start_color='FFE0A960', end_color='FFE0A960', fill_type='solid')
    # Definir los estilos para gris
    gray_fill = PatternFill(start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid')
    # Definir el estilo de color verde
    green_fill = PatternFill(start_color='FF60E0B3', end_color='FF60E0B3', fill_type='solid')
    # Definir el estilo de color amarillo
    yellow_fill = PatternFill(start_color='FFE0DE60', end_color='FFE0DE60', fill_type='solid')
    # Definir el estilo de color azul
    blue_fill = PatternFill(start_color='FF60A2E0', end_color='FF60A2E0', fill_type='solid')
    # Definir el estilo de color verde 2
    green_fill_2 = PatternFill(start_color='FF60E07E', end_color='FF60E07E', fill_type='solid')   
    
    green_font = Font(name='Arial', size=8, color='00FF00')  # Verde
    red_font = Font(name='Arial', size=8, color='FF0000')    # Rojo
    
    border = Border(left=Side(style='thin', color='00B0F0'),
                    right=Side(style='thin', color='00B0F0'),
                    top=Side(style='thin', color='00B0F0'),
                    bottom=Side(style='thin', color='00B0F0'))
    borde_plomo = Border(left=Side(style='thin', color='A9A9A9'), # Plomo
                    right=Side(style='thin', color='A9A9A9'), # Plomo
                    top=Side(style='thin', color='A9A9A9'), # Plomo
                    bottom=Side(style='thin', color='A9A9A9')) # Plomo
    
    # Definir el estilo de relleno celeste
    celeste_fill = PatternFill(start_color='FF87CEEB', end_color='FF87CEEB', fill_type='solid')
    # Morado más claro
    morado_claro_fill = PatternFill(start_color='FFE9D8FF', end_color='FFE9D8FF', fill_type='solid')
    # Plomo más claro
    plomo_claro_fill = PatternFill(start_color='FFEDEDED', end_color='FFEDEDED', fill_type='solid')
    # Azul más claro
    azul_claro_fill = PatternFill(start_color='FFD8EFFA', end_color='FFD8EFFA', fill_type='solid')
    # Naranja más claro
    naranja_claro_fill = PatternFill(start_color='FFFFEBD8', end_color='FFFFEBD8', fill_type='solid')
    # Verde más claro
    verde_claro_fill = PatternFill(start_color='FFBDF7BD', end_color='FFBDF7BD', fill_type='solid')
    
    border_negro = Border(left=Side(style='thin', color='000000'), # negro
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'), 
        bottom=Side(style='thin', color='000000')) 
    
        ### CAMBIO DE CABECERAS     
    ####################################
    # Merge cells 
    # numerador y denominador
    ws.merge_cells('B5:J6') 
    ws.merge_cells('K5:AN5')
    
    # BPN y Sin BPN
    ws.merge_cells('K6:Q6') 
    ws.merge_cells('R6:AF6')
    ws.merge_cells('AG6:AN6')
    
    # Auxiliar HORIZONTAL
    ws.merge_cells('Q7:Q8')
    ws.merge_cells('AF7:AF8')
    ws.merge_cells('AK7:AK8')
    ws.merge_cells('AL7:AL8')
    ws.merge_cells('AM7:AM8')
    ws.merge_cells('AN7:AN8')
    
    # intervalo
    ws.merge_cells('B7:C7')
    ws.merge_cells('D7:F7')
    ws.merge_cells('G7:H7')
    ws.merge_cells('I7:J7')
    ws.merge_cells('K7:L7')
    ws.merge_cells('M7:N7')
    ws.merge_cells('O7:P7')
    ws.merge_cells('R7:S7')
    ws.merge_cells('T7:U7')
    ws.merge_cells('V7:W7')
    ws.merge_cells('X7:Y7')
    ws.merge_cells('Z7:AA7')
    ws.merge_cells('AB7:AC7')
    ws.merge_cells('AD7:AE7')
    ws.merge_cells('AG7:AH7')
    ws.merge_cells('AI7:AJ7')
    
    
    # COD HIS
    ws.merge_cells('B8:F8')
    ws.merge_cells('G8:H8')
    ws.merge_cells('I8:J8')
    ws.merge_cells('K8:L8')
    ws.merge_cells('M8:N8')
    ws.merge_cells('O8:P8')
    ws.merge_cells('R8:S8')
    ws.merge_cells('T8:U8')
    ws.merge_cells('V8:AE8')
    ws.merge_cells('AG8:AH8')
    ws.merge_cells('AI8:AJ8')
    
    # Combina cela
    ws['B5'] = 'DENOMINADOR'
    ws['K5'] = 'NUMERADOR'
    
    ws['K6'] = 'ESQUEMA DE SUPLEMENTACION TRATAMIENTO CON HIERRO'
    ws['R6'] = 'ESQUEMA DE SUPLEMENTACION TRATAMIENTO CON MULTIMICRONUTIENTES'
    ws['AG6'] = 'VALIDADORES DEL INDICADOR'
    
    ws['D7'] = 'Niños que cumplan 12 meses 29 dias (394) dentro del mes de evaluacion'
    ws['G7'] = 'Dosaje de hemoglobina a los 6 meses de edad (Se busca el registro entre 170 a 209 dias)'
    ws['I7'] = 'SIN Dx de anemia a los 6 meses (170-209 días) tras el primer dosaje de hemoglobina'
    ws['K7'] = '1° Entrega de hierro el mismo dia del 1er Dosaje de Hemoglobina'
    ws['M7'] = 'Al tercer mes de iniciada la suplementacion (Se busca el registro entre 90 - 119 días)'
    ws['O7'] = '2° Entrega de hierro tenga un intervalo entre 25 a 70 dias respecto a la 1° Entrega'
    ws['Q7'] = 'Cumple con el Termino de la suplmentacion TA'
    ws['R7'] = '1° Entrega de hierro el mismo dia del 1er Dosaje de Hemoglobina'
    ws['T7'] = 'Al tercer mes de iniciada la suplementacion (Se busca el registro entre 90 - 119 días)'
    ws['V7'] = '2° Entrega de hierro tenga un intervalo entre 25 a 70 dias respecto a la 1° Entrega'
    ws['X7'] = '3° Entrega de hierro tenga un intervalo entre 25 a 70 dias respecto a la 2° Entregac'
    ws['Z7'] = '4° Entrega de hierro tenga un intervalo entre 25 a 70 dias respecto a la 3° Entrega'
    ws['AB7'] = '5° Entrega de hierro tenga un intervalo entre 25 a 70 dias respecto a la 4° Entrega'
    ws['AD7'] = '6° Entrega de hierro tenga un intervalo entre 25 a 70 dias respecto a la 5° Entrega'
    ws['AF7'] = 'Cumple con el Termino de la suplmentacion TA'    
    ws['AG7'] = 'Registrado a los seis (06) meses de iniciado el tratamiento (se busca el registro entre 170-209 dias)'
    ws['AI7'] = 'Dosaje de Hemoglobina a los 12 meses de edad (se busca el dato entre entre 365 días a 394 días)'
    ws['AK7'] = 'Reciben una (01) entrega de Hierro o una (01) entrega de multimicronutrientes '
    ws['AL7'] = 'Reciben un dosaje de hemoglobina de control al tercer mes de iniciada la suplementaciónc'
    ws['AM7'] = 'Reciben un dosaje de hemoglobina de control al tercer mes de iniciada la suplementación'
    ws['AN7'] = 'Cumple con el Termino de la suplmentacion TA'
    
    ws['G8'] = 'DX = 85018 ó 85018.01 + TD = D'
    ws['I8'] = 'DX = D500 ó D508 ó D509 ó D649 ó D539 + TD =D'    
    ws['K8'] = 'DX = 99199.17'
    ws['M8'] = 'DX = 85018 ó 85018.01 + TD = D'
    ws['O8'] = 'DX = 99199.17'
    ws['R8'] = 'DX = 99199.19'
    ws['T8'] = 'DX = 85018 ó 85018.01 + TD = D'
    ws['V8'] = 'DX = 99199.19'
    ws['AG8'] = 'DX = 99199.17 ó 99199.19 + LAB = TA'
    ws['AI8'] = 'DX = 85018 ó 85018.01 + TD = D'
    
    ### numerador y denominador 
    
    ws['B5'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['B5'].font = Font(name = 'Arial', size= 10, bold = True)
    ws['B5'].fill = gray_fill
    ws['B5'].border = border_negro
    
    ws['K5'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['K5'].font = Font(name = 'Arial', size= 10, bold = True)
    ws['K5'].fill = naranja_claro_fill
    ws['K5'].border = border_negro
    
    ### SUPLEMENTACION HIERRO Y MULTIMICRONUTIENTES 
    
    ws['K6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['K6'].font = Font(name = 'Arial', size= 9, bold = True)
    ws['K6'].fill = morado_claro_fill
    ws['K6'].border = border_negro
    
    ws['R6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['R6'].font = Font(name = 'Arial', size= 9, bold = True)
    ws['R6'].fill = verde_claro_fill
    ws['R6'].border = border_negro
    
    ws['AG6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AG6'].font = Font(name = 'Arial', size= 9, bold = True)
    ws['AG6'].fill = naranja_claro_fill
    ws['AG6'].border = border_negro
    
    
    ### intervalo 
    ws['D7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['D7'].font = Font(name = 'Arial', size= 7)
    ws['D7'].fill = plomo_claro_fill
    ws['D7'].border = border_negro
    
    ws['G7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['G7'].font = Font(name = 'Arial', size= 7)
    ws['G7'].fill = plomo_claro_fill
    ws['G7'].border = border_negro

    ws['I7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['I7'].font = Font(name = 'Arial', size= 7)
    ws['I7'].fill = plomo_claro_fill
    ws['I7'].border = border_negro
    
    ws['K7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['K7'].font = Font(name = 'Arial', size= 7)
    ws['K7'].fill = plomo_claro_fill
    ws['K7'].border = border_negro    
    
    ws['M7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['M7'].font = Font(name = 'Arial', size= 7)
    ws['M7'].fill = plomo_claro_fill
    ws['M7'].border = border_negro
    
    ws['N7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['N7'].font = Font(name = 'Arial', size= 7)
    ws['N7'].fill = plomo_claro_fill
    ws['N7'].border = border_negro
    
    ws['O7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['O7'].font = Font(name = 'Arial', size= 7)
    ws['O7'].fill = plomo_claro_fill
    ws['O7'].border = border_negro
    
    ws['Q7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['Q7'].font = Font(name = 'Arial', size= 7)
    ws['Q7'].fill = plomo_claro_fill
    ws['Q7'].border = border_negro
    
    ws['R7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['R7'].font = Font(name = 'Arial', size= 7)
    ws['R7'].fill = plomo_claro_fill
    ws['R7'].border = border_negro
    
    ws['T7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['T7'].font = Font(name = 'Arial', size= 7)
    ws['T7'].fill = plomo_claro_fill
    ws['T7'].border = border_negro
    
    ws['V7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['V7'].font = Font(name = 'Arial', size= 7)
    ws['V7'].fill = plomo_claro_fill
    ws['V7'].border = border_negro
    
    ws['X7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['X7'].font = Font(name = 'Arial', size= 7)
    ws['X7'].fill = plomo_claro_fill
    ws['X7'].border = border_negro
    
    ws['Z7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['Z7'].font = Font(name = 'Arial', size= 7)
    ws['Z7'].fill = plomo_claro_fill
    ws['Z7'].border = border_negro
    
    ws['AB7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AB7'].font = Font(name = 'Arial', size= 7)
    ws['AB7'].fill = plomo_claro_fill
    ws['AB7'].border = border_negro
    
    ws['AD7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AD7'].font = Font(name = 'Arial', size= 7)
    ws['AD7'].fill = plomo_claro_fill
    ws['AD7'].border = border_negro
    
    ws['AF7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AF7'].font = Font(name = 'Arial', size= 7)
    ws['AF7'].fill = plomo_claro_fill
    ws['AF7'].border = border_negro
    
    ws['AG7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AG7'].font = Font(name = 'Arial', size= 7)
    ws['AG7'].fill = plomo_claro_fill
    ws['AG7'].border = border_negro
    
    ws['AI7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AI7'].font = Font(name = 'Arial', size= 7)
    ws['AI7'].fill = plomo_claro_fill
    ws['AI7'].border = border_negro
    
    ws['AK7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AK7'].font = Font(name = 'Arial', size= 7)
    ws['AK7'].fill = plomo_claro_fill
    ws['AK7'].border = border_negro
    
    ws['AL7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AL7'].font = Font(name = 'Arial', size= 7)
    ws['AL7'].fill = plomo_claro_fill
    ws['AL7'].border = border_negro
    
    ws['AM7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AM7'].font = Font(name = 'Arial', size= 7)
    ws['AM7'].fill = plomo_claro_fill
    ws['AM7'].border = border_negro
    
    ws['AN7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AN7'].font = Font(name = 'Arial', size= 7)
    ws['AN7'].fill = plomo_claro_fill
    ws['AN7'].border = border_negro
    
    ws['B7'].alignment = Alignment(horizontal= "center", vertical="center")
    ws['B7'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['B7'].fill = plomo_claro_fill
    ws['B7'].border = border_negro
    ws['B7'] = 'INTERVALO'
    
    ws['B8'].alignment = Alignment(horizontal= "center", vertical="center")
    ws['B8'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['B8'].fill = azul_claro_fill
    ws['B8'].border = border_negro
    ws['B8'] = 'COD HIS'
    
    ### cod_his 
    ws['G8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['G8'].font = Font(name = 'Arial', size= 7)
    ws['G8'].fill = azul_claro_fill
    ws['G8'].border = border_negro

    ws['I8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['I8'].font = Font(name = 'Arial', size= 7)
    ws['I8'].fill = azul_claro_fill
    ws['I8'].border = border_negro
    
    ws['K8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['K8'].font = Font(name = 'Arial', size= 7)
    ws['K8'].fill = azul_claro_fill
    ws['K8'].border = border_negro    
    
    ws['M8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['M8'].font = Font(name = 'Arial', size= 7)
    ws['M8'].fill = azul_claro_fill
    ws['M8'].border = border_negro
    
    ws['O8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['O8'].font = Font(name = 'Arial', size= 7)
    ws['O8'].fill = azul_claro_fill
    ws['O8'].border = border_negro
    
    ws['R8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['R8'].font = Font(name = 'Arial', size= 7)
    ws['R8'].fill = azul_claro_fill
    ws['R8'].border = border_negro
    
    ws['T8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['T8'].font = Font(name = 'Arial', size= 7)
    ws['T8'].fill = azul_claro_fill
    ws['T8'].border = border_negro
    
    ws['V8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['V8'].font = Font(name = 'Arial', size= 7)
    ws['V8'].fill = azul_claro_fill
    ws['V8'].border = border_negro
    
    ws['AG8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AG8'].font = Font(name = 'Arial', size= 7)
    ws['AG8'].fill = azul_claro_fill
    ws['AG8'].border = border_negro
    
    ws['AI8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AI8'].font = Font(name = 'Arial', size= 7)
    ws['AI8'].fill = azul_claro_fill
    ws['AI8'].border = border_negro    
    ### BORDE DE CELDAS CONBINADAS
    
    # NUM y DEN
    inicio_columna = 'B'
    fin_columna = 'AN'
    fila = 5
    from openpyxl.utils import column_index_from_string
    # Convertir letras de columna a índices numéricos
    indice_inicio = column_index_from_string(inicio_columna)
    indice_fin = column_index_from_string(fin_columna)
    # Iterar sobre las columnas en el rango especificado
    for col in range(indice_inicio, indice_fin + 1):
        celda = ws.cell(row=fila, column=col)
        celda.border = border_negro
    
    # INTERVALO
    inicio_columna = 'K'
    fin_columna = 'AN'
    fila = 6
    from openpyxl.utils import column_index_from_string
    # Convertir letras de columna a índices numéricos
    indice_inicio = column_index_from_string(inicio_columna)
    indice_fin = column_index_from_string(fin_columna)
    # Iterar sobre las columnas en el rango especificado
    for col in range(indice_inicio, indice_fin + 1):
        celda = ws.cell(row=fila, column=col)
        celda.border = border_negro
        
    # CODIGO HIS 
    inicio_columna = 'B'
    fin_columna = 'AN'
    fila = 7
    from openpyxl.utils import column_index_from_string
    # Convertir letras de columna a índices numéricos
    indice_inicio = column_index_from_string(inicio_columna)
    indice_fin = column_index_from_string(fin_columna)
    # Iterar sobre las columnas en el rango especificado
    for col in range(indice_inicio, indice_fin + 1):
        celda = ws.cell(row=fila, column=col)
        celda.border = border_negro
    
    # CODIGO HIS 
    inicio_columna = 'B'
    fin_columna = 'AN'
    fila = 8
    from openpyxl.utils import column_index_from_string
    # Convertir letras de columna a índices numéricos
    indice_inicio = column_index_from_string(inicio_columna)
    indice_fin = column_index_from_string(fin_columna)
    # Iterar sobre las columnas en el rango especificado
    for col in range(indice_inicio, indice_fin + 1):
        celda = ws.cell(row=fila, column=col)
        celda.border = border_negro
    
    ##### imprimer fecha y hora del reporte
    fecha_hora_actual = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    nombre_usuario = getpass.getuser()

    # Obtener el usuario actualmente autenticado
    try:
        user = User.objects.get(is_active=True)
    except User.DoesNotExist:
        user = None
    except User.MultipleObjectsReturned:
        # Manejar el caso donde hay múltiples usuarios activos
        user = User.objects.filter(is_active=True).first()  # Por ejemplo, obtener el primero
    # Asignar fecha y hora a la celda A1
    ws['V1'].value = 'Fecha y Hora:'
    ws['W1'].value = fecha_hora_actual

    # Asignar nombre de usuario a la celda A2
    ws['V2'].value = 'Usuario:'
    ws['W2'].value = nombre_usuario
    
    # Formatear las etiquetas en negrita
    etiqueta_font = Font(name='Arial', size=8)
    ws['V1'].font = etiqueta_font
    ws['W1'].font = etiqueta_font
    ws['V2'].font = etiqueta_font
    ws['W2'].font = etiqueta_font

    # Alinear el texto
    ws['V1'].alignment = Alignment(horizontal="right", vertical="center")
    ws['W1'].alignment = Alignment(horizontal="left", vertical="center")
    ws['V2'].alignment = Alignment(horizontal="right", vertical="center")
    ws['W2'].alignment = Alignment(horizontal="left", vertical="center")
        
    ## crea titulo del reporte
    ws['B1'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B1'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['B1'] = 'OFICINA DE TECNOLOGIAS DE LA INFORMACION'
    
    ws['B2'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B2'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['B2'] = 'DIRECCION REGIONAL DE SALUD JUNIN'
    
    ws['B4'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B4'].font = Font(name = 'Arial', size= 12, bold = True)
    ws['B4'] = 'SEGUIMIENTO NOMINAL DEL INDICADOR SI-02.03. PORCENTAJE DE NIÑAS Y NIÑOS DE DOCE (12) MESES DE EDAD, SIN DIAGNÓSTICO DE ANEMIA, QUE RECIBEN TRES (03) DOSAJES DE HEMOGLOBINA Y CULMINAN LA SUPLEMENTACIÓN PREVENTIVA CON HIERRO'
    
    ws['B3'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B3'].font = Font(name = 'Arial', size= 7, bold = True, color='0000CC')
    ws['B3'] ='El usuario se compromete a mantener la confidencialidad de los datos personales que conozca como resultado del reporte realizado, cumpliendo con lo establecido en la Ley N° 29733 - Ley de Protección de Datos Personales y sus normas complementarias.'
        
    ws['B9'].alignment = Alignment(horizontal= "center", vertical="center")
    ws['B9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['B9'].fill = fill
    ws['B9'].border = border
    ws['B9'] = 'TD'
        
    ws['C9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['C9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['C9'].fill = fill
    ws['C9'].border = border
    ws['C9'] = 'NUM DOC'      
    
    ws['D9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['D9'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['D9'].fill = fill
    ws['D9'].border = border
    ws['D9'] = 'FECHA NAC' 
    
    ws['E9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['E9'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['E9'].fill = fill
    ws['E9'].border = border
    ws['E9'] = 'SEXO'     
    
    ws['F9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['F9'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['F9'].fill = fill
    ws['F9'].border = border
    ws['F9'] = 'SEGURO'    
    
    ws['G9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['G9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['G9'].fill = fill
    ws['G9'].border = border
    ws['G9'] = 'HB'    
    
    ws['H9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['H9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['H9'].fill = fill
    ws['H9'].border = border
    ws['H9'] = 'VAL HB'  
    
    ws['I9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['I9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['I9'].fill = fill
    ws['I9'].border = border
    ws['I9'] = 'DX ANEMIA'    
    
    ws['J9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['J9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['J9'].fill = fill
    ws['J9'].border = border
    ws['J9'] = 'VAL'  
    
    ws['K9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['K9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['K9'].fill = green_fill
    ws['K9'].border = border
    ws['K9'] = '1° SUP'    
    
    ws['L9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['L9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['L9'].fill = green_fill
    ws['L9'].border = border
    ws['L9'] = 'VAL' 
    
    ws['M9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['M9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['M9'].fill = green_fill_2
    ws['M9'].border = border
    ws['M9'] = 'HB CONTROL' 
    
    ws['N9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['N9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['N9'].fill = green_fill_2
    ws['N9'].border = border
    ws['N9'] = 'VAL' 
    
    ws['O9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['O9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['O9'].fill = green_fill
    ws['O9'].border = border
    ws['O9'] = '2° SUP' 
    
    ws['P9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['P9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['P9'].fill = green_fill
    ws['P9'].border = border
    ws['P9'] = 'VAL' 
    
    ws['Q9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['Q9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['Q9'].fill = green_fill
    ws['Q9'].border = border
    ws['Q9'] = 'TA SUP PO' 
    
    ws['R9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['R9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['R9'].fill = green_fill
    ws['R9'].border = border
    ws['R9'] = '1° SUP MULT' 
    
    ws['S9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['S9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['S9'].fill = green_fill
    ws['S9'].border = border
    ws['S9'] = 'VAL' 
    
    ws['T9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['T9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['T9'].fill = green_fill_2
    ws['T9'].border = border
    ws['T9'] = 'HB CONTROL' 
    
    ws['U9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['U9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['U9'].fill = green_fill_2
    ws['U9'].border = border
    ws['U9'] = 'VAL' 
    
    ws['V9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['V9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['V9'].fill = green_fill
    ws['V9'].border = border
    ws['V9'] = '2° SUP MULT' 
    
    ws['W9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['W9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['W9'].fill = green_fill
    ws['W9'].border = border
    ws['W9'] = 'VAL' 
    
    ws['X9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['X9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['X9'].fill = green_fill
    ws['X9'].border = border
    ws['X9'] = '3° SUP MULT' 
    
    ws['Y9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['Y9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['Y9'].fill = green_fill
    ws['Y9'].border = border
    ws['Y9'] = 'VAL' 
    
    ws['Z9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['Z9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['Z9'].fill = green_fill
    ws['Z9'].border = border
    ws['Z9'] = '4° SUP MULT' 
    
    ws['AA9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AA9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['AA9'].fill = green_fill
    ws['AA9'].border = border
    ws['AA9'] = 'VAL' 
    
    ws['AB9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AB9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['AB9'].fill = green_fill
    ws['AB9'].border = border
    ws['AB9'] = '5° SUP MULT' 
    
    ws['AC9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AC9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['AC9'].fill = green_fill
    ws['AC9'].border = border
    ws['AC9'] = 'VAL' 
    
    ws['AD9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AD9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['AD9'].fill = green_fill
    ws['AD9'].border = border
    ws['AD9'] = '6° SUP MULT' 
    
    ws['AE9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AE9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['AE9'].fill = green_fill
    ws['AE9'].border = border
    ws['AE9'] = 'VAL' 
    
    ws['AF9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AF9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['AF9'].fill = green_fill
    ws['AF9'].border = border
    ws['AF9'] = 'SUP MUL TA' 
    
    ws['AG9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AG9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['AG9'].fill = green_fill
    ws['AG9'].border = border
    ws['AG9'] = 'TA' 
    
    ws['AH9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AH9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['AH9'].fill = green_fill
    ws['AH9'].border = border
    ws['AH9'] = 'IND TA' 
    
    ws['AI9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AI9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['AI9'].fill = green_fill_2
    ws['AI9'].border = border
    ws['AI9'] = 'HB 12M' 
    
    ws['AJ9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AJ9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['AJ9'].fill = green_fill_2
    ws['AJ9'].border = border
    ws['AJ9'] = 'IND HB 12M' 
    
    ws['AK9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AK9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['AK9'].fill = naranja_claro_fill
    ws['AK9'].border = border
    ws['AK9'] = 'VAL 1° SUP' 
    
    ws['AL9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AL9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['AL9'].fill = naranja_claro_fill
    ws['AL9'].border = border
    ws['AL9'] = 'SUP CONT' 
    
    ws['AM9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AM9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['AM9'].fill = naranja_claro_fill
    ws['AM9'].border = border
    ws['AM9'] = 'SUP CONTC' 
    
    ws['AN9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AN9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['AN9'].fill = naranja_claro_fill
    ws['AN9'].border = border
    ws['AN9'] = 'SUP TA' 
    
    ws['AO9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AO9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['AO9'].fill = fill
    ws['AO9'].border = border
    ws['AO9'] = 'MES' 
    
    ws['AP9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AP9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['AP9'].fill = gray_fill
    ws['AP9'].border = border
    ws['AP9'] = 'IND' 
    
    ws['AQ9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AQ9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['AQ9'].fill = orange_fill
    ws['AQ9'].border = border
    ws['AQ9'] = 'UBIGEO'  
    
    ws['AR9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AR9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['AR9'].fill = orange_fill
    ws['AR9'].border = border
    ws['AR9'] = 'PROVINCIA'       
    
    ws['AS9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AS9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['AS9'].fill = orange_fill
    ws['AS9'].border = border
    ws['AS9'] = 'DISTRITO' 
    
    ws['AT9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AT9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['AT9'].fill = orange_fill
    ws['AT9'].border = border
    ws['AT9'] = 'RED'  
    
    ws['AU9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AU9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['AU9'].fill = orange_fill
    ws['AU9'].border = border
    ws['AU9'] = 'MICRORED'  
    
    ws['AV9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AV9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['AV9'].fill = orange_fill
    ws['AV9'].border = border
    ws['AV9'] = 'COD EST'  
    
    ws['AW9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AW9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['AW9'].fill = orange_fill
    ws['AW9'].border = border
    ws['AW9'] = 'ESTABLECIMIENTO'  
    
    # Definir estilos
    header_font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    centered_alignment = Alignment(horizontal='center')
    border = Border(left=Side(style='thin', color='A9A9A9'),
            right=Side(style='thin', color='A9A9A9'),
            top=Side(style='thin', color='A9A9A9'),
            bottom=Side(style='thin', color='A9A9A9'))
    header_fill = PatternFill(patternType='solid', fgColor='00B0F0')
    
    
    # Definir los caracteres especiales de check y X
    check_mark = '✓'  # Unicode para check
    x_mark = '✗'  # Unicode para X
    sub_cumple = 'CUMPLE'
    sub_no_cumple = 'NO CUMPLE'
    
    # Define styles
    promo_fill = PatternFill(patternType='solid', fgColor='FFD966')  # Yellow fill for promo
    font_normal = Font(name='Arial', size=8)
    font_bold_white = Font(name='Arial', size=7, bold=True, color='FFFFFF')
    font_red_bold = Font(name='Arial', size=7, bold=True, color='FF0000')
    font_green_bold = Font(name='Arial', size=7, bold=True, color='00FF00')
    font_red = Font(name='Arial', size=7, color='FF0000')
    font_green = Font(name='Arial', size=7, color='00B050')
    font_check = Font(name='Arial', size=10, color='00B050')
    font_x = Font(name='Arial', size=10, color='FF0000')
    plomo_claro_font = Font(name='Arial', size=7, color='FFEDEDED', bold=False)

    # Define fills
    fill_red = PatternFill(patternType='solid', fgColor='FF0000')
    fill_green = PatternFill(patternType='solid', fgColor='00FF00')
    
    # Write data
    for row, record in enumerate(results, start=10):
        bpn_value = record[8]
        multi_value = record[10]# BPN is in column 6 (index 4)
        for col_offset, value in enumerate(record):
            col = col_offset + 2  # Adjust column index (starts at 2)
            cell = ws.cell(row=row, column=col, value=value)

            # Alignment
            if col in [28, 29, 32]:
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='center')

            # Initialize default font and fill
            cell_font = font_normal
            cell_fill = None

            # Apply special formatting based on column and value
            if col == 42:
                if isinstance(value, str):
                    value_upper = value.strip().upper()
                    if value_upper == "NO CUMPLE":
                        cell_fill = fill_red
                        cell_font = font_bold_white
                    elif value_upper == "CUMPLE":
                        cell_fill = fill_green
                        cell_font = font_bold_white
                    else:
                        cell_font = Font(name='Arial', size=7)
                else:
                    cell_font = font_normal

            elif col in [17, 32, 37, 38, 39, 40]:
                if value == 0:
                    cell.value = sub_no_cumple
                    cell_font = font_red
                elif value == 1:
                    cell.value = sub_cumple
                    cell_font = font_green
                else:
                    cell_font = Font(name='Arial', size=7)

            elif col in [8,10,12,14,16,19,21,23,25,27,29,31,34,36]:
                if value == 1:
                    cell.value = check_mark
                    cell_font = font_check
                elif value == 0:
                    cell.value = x_mark
                    cell_font = font_x
                else:
                    cell_font = font_normal

            # Apply color to columns H (8) to Q (17) and write "NO APLICA" in column R (18) if BPN is 0
            if bpn_value == 0:
                if 9 <= col <= 10:
                    cell_fill = plomo_claro_fill
                    cell_font = plomo_claro_font
                elif col == 9:
                    cell.value = "NO APLICA"
                    cell_fill = plomo_claro_fill
                    cell_font = font_normal

            # Apply color to columns H (8) to Q (17) and write "NO APLICA" in column R (18) if BPN is 0
            if multi_value == 1:
                if 18 <= col <= 31:
                    cell_fill = plomo_claro_fill
                    cell_font = plomo_claro_font
                elif col == 32:
                    cell.value = "NO APLICA"
                    cell_fill = plomo_claro_fill
                    cell_font = font_normal
            
            
            # Set font and fill
            cell.font = cell_font
            if cell_fill:
                cell.fill = cell_fill

            # Apply borders
            cell.border = border


###########################################################################################
# -- COBERTURA PAQUETE NEONATAL
def obtener_cobertura_s23_suplementacion12():
    with connection.cursor() as cursor:
        cursor.execute(
            'SELECT * FROM public."Cobertura_SI_0203_Suplementacion12M" ORDER BY "Red", "MicroRed", "Nombre_Establecimiento";'
        )
        return cursor.fetchall()

class RptCoberturaS23Suplementacion12(TemplateView):
    def get(self, request, *args, **kwargs):
        # Variables ingresadas
                
        # Creación de la consulta
        resultado_cobertura = obtener_cobertura_s23_suplementacion12()
        
        wb = Workbook()
        
        consultas = [
                ('Cobertura', resultado_cobertura)
        ]
        
        for index, (sheet_name, results) in enumerate(consultas):
            if index == 0:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
        
            fill_worksheet_cobertura_s23_suplementacion12(ws, results)
        
        ##########################################################################          
        # Establecer el nombre del archivo
        nombre_archivo = "rpt_cobertura_s23_suplementacion12.xlsx"
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)

        return response

def fill_worksheet_cobertura_s23_suplementacion12(ws, results): 
    # cambia el alto de la columna
    ws.row_dimensions[1].height = 14
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 3
    ws.row_dimensions[4].height = 25
    ws.row_dimensions[5].height = 3
    ws.row_dimensions[7].height = 20
    ws.row_dimensions[8].height = 130
    # cambia el ancho de la columna
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 8
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 8
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 8
    ws.column_dimensions['N'].width = 15
    ws.column_dimensions['O'].width = 15
    ws.column_dimensions['P'].width = 8
    ws.column_dimensions['Q'].width = 15
    ws.column_dimensions['R'].width = 15
    ws.column_dimensions['S'].width = 8
    ws.column_dimensions['T'].width = 15
    ws.column_dimensions['U'].width = 15
    ws.column_dimensions['V'].width = 8
    ws.column_dimensions['W'].width = 15
    ws.column_dimensions['X'].width = 15
    ws.column_dimensions['Y'].width = 8
    ws.column_dimensions['Z'].width = 15    
    ws.column_dimensions['AA'].width = 15
    ws.column_dimensions['AB'].width = 8       
    ws.column_dimensions['AC'].width = 15
    ws.column_dimensions['AD'].width = 15    
    ws.column_dimensions['AE'].width = 8
    ws.column_dimensions['AF'].width = 15    
    ws.column_dimensions['AG'].width = 15
    ws.column_dimensions['AH'].width = 8    
    ws.column_dimensions['AI'].width = 15
    ws.column_dimensions['AJ'].width = 15    
    ws.column_dimensions['AK'].width = 8
    ws.column_dimensions['AL'].width = 15    
    ws.column_dimensions['AM'].width = 15
    ws.column_dimensions['AN'].width = 8    
    ws.column_dimensions['AO'].width = 15
    ws.column_dimensions['AP'].width = 15    
    ws.column_dimensions['AQ'].width = 8    
    
    # linea de division
    ws.freeze_panes = 'H9'
    # Configuración del fondo y el borde
    # Definir el color usando formato aRGB (opacidad completa 'FF' + color RGB)
    fill = PatternFill(start_color='FF60D7E0', end_color='FF60D7E0', fill_type='solid')
    # Definir el color anaranjado usando formato aRGB
    orange_fill = PatternFill(start_color='FFE0A960', end_color='FFE0A960', fill_type='solid')
    # Definir los estilos para gris
    gray_fill = PatternFill(start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid')
    # Definir el estilo de color verde
    green_fill = PatternFill(start_color='FF60E0B3', end_color='FF60E0B3', fill_type='solid')
    # Definir el estilo de color amarillo
    yellow_fill = PatternFill(start_color='FFE0DE60', end_color='FFE0DE60', fill_type='solid')
    # Definir el estilo de color azul
    blue_fill = PatternFill(start_color='FF60A2E0', end_color='FF60A2E0', fill_type='solid')
    # Definir el estilo de color verde 2
    green_fill_2 = PatternFill(start_color='FF60E07E', end_color='FF60E07E', fill_type='solid')
    
    green_font = Font(name='Arial', size=8, color='00FF00')  # Verde
    red_font = Font(name='Arial', size=8, color='FF0000')    # Rojo
    
    border = Border(left=Side(style='thin', color='00B0F0'),
                    right=Side(style='thin', color='00B0F0'),
                    top=Side(style='thin', color='00B0F0'),
                    bottom=Side(style='thin', color='00B0F0'))
    
    borde_plomo = Border(left=Side(style='thin', color='A9A9A9'), # Plomo
                    right=Side(style='thin', color='A9A9A9'), # Plomo
                    top=Side(style='thin', color='A9A9A9'), # Plomo
                    bottom=Side(style='thin', color='A9A9A9')) # Plomo
    
    border_negro = Border(left=Side(style='thin', color='000000'), # negro
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'), 
                    bottom=Side(style='thin', color='000000')) 
    
    # Merge cells 
    ws.merge_cells('E7:G7') 
    ws.merge_cells('H7:J7')
    ws.merge_cells('K7:M7')
    ws.merge_cells('N7:P7')
    ws.merge_cells('Q7:S7')
    ws.merge_cells('T7:V7')
    ws.merge_cells('W7:Y7')
    ws.merge_cells('Z7:AB7')
    ws.merge_cells('AC7:AE7')
    ws.merge_cells('AF7:AH7')
    ws.merge_cells('AI7:AK7')
    ws.merge_cells('AL7:AN7')
    ws.merge_cells('AO7:AQ7')

    # Set the value for the merged cell
    ws['E7'] = 'TOTAL'
    ws['H7'] = 'ENERO'
    ws['K7'] = 'FEBRERO'
    ws['N7'] = 'MARZO'
    ws['Q7'] = 'ABRIL'
    ws['T7'] = 'MAYO'
    ws['W7'] = 'JUNIO'
    ws['Z7'] = 'JULIO'
    ws['AC7'] = 'AGOSTO'
    ws['AF7'] = 'SETIEMBRE'
    ws['AI7'] = 'OCTUBRE'
    ws['AL7'] = 'NOVIEMBRE'
    ws['AO7'] = 'DICIEMBRE'

    # Definir el rango desde B3 hasta AA3
    inicio_columna = 'E'
    fin_columna = 'AQ'
    fila = 7
    from openpyxl.utils import column_index_from_string
    # Convertir letras de columna a índices numéricos
    indice_inicio = column_index_from_string(inicio_columna)
    indice_fin = column_index_from_string(fin_columna)

    # Iterar sobre las columnas en el rango especificado
    for col in range(indice_inicio, indice_fin + 1):
        celda = ws.cell(row=fila, column=col)
        celda.border = border_negro
    
    # Definir el rango desde B3 hasta AA3
    inicio_columna_cab = 'B'
    fin_columna_cab = 'AQ'
    fila = 8
    
    # Convertir letras de columna a índices numéricos
    indice_inicio_cab = column_index_from_string(inicio_columna_cab)
    indice_fin_cab = column_index_from_string(fin_columna_cab)

    # Iterar sobre las columnas en el rango especificado
    for col in range(indice_inicio_cab, indice_fin_cab + 1):
        celda = ws.cell(row=fila, column=col)
        celda.border = border_negro
    
    # Apply formatting to the merged cell
    ws['E7'].alignment = Alignment(horizontal="center", vertical="center")
    ws['E7'].font = Font(name='Arial', size=8, bold=True, color='000000')
    ws['E7'].fill = yellow_fill  # Assuming yellow_fill is predefined
    ws['E7'].border = border_negro     # Assuming border is predefined

    ws['H7'].alignment = Alignment(horizontal="center", vertical="center")
    ws['H7'].font = Font(name='Arial', size=8, bold=True, color='000000')
    ws['H7'].fill = blue_fill  # Assuming yellow_fill is predefined
    ws['H7'].border = border_negro 
    
    ws['K7'].alignment = Alignment(horizontal="center", vertical="center")
    ws['K7'].font = Font(name='Arial', size=8, bold=True, color='000000')
    ws['K7'].fill = blue_fill  # Assuming yellow_fill is predefined
    ws['K7'].border = border_negro 
    
    ws['N7'].alignment = Alignment(horizontal="center", vertical="center")
    ws['N7'].font = Font(name='Arial', size=8, bold=True, color='000000')
    ws['N7'].fill = blue_fill  # Assuming yellow_fill is predefined
    ws['N7'].border = border_negro 
    
    ws['Q7'].alignment = Alignment(horizontal="center", vertical="center")
    ws['Q7'].font = Font(name='Arial', size=8, bold=True, color='000000')
    ws['Q7'].fill = blue_fill  # Assuming yellow_fill is predefined
    ws['Q7'].border = border_negro 
    
    ws['T7'].alignment = Alignment(horizontal="center", vertical="center")
    ws['T7'].font = Font(name='Arial', size=8, bold=True, color='000000')
    ws['T7'].fill = blue_fill  # Assuming yellow_fill is predefined
    ws['T7'].border = border_negro 
    
    ws['W7'].alignment = Alignment(horizontal="center", vertical="center")
    ws['W7'].font = Font(name='Arial', size=8, bold=True, color='000000')
    ws['W7'].fill = blue_fill  # Assuming yellow_fill is predefined
    ws['W7'].border = border_negro 
    
    ws['Z7'].alignment = Alignment(horizontal="center", vertical="center")
    ws['Z7'].font = Font(name='Arial', size=8, bold=True, color='000000')
    ws['Z7'].fill = blue_fill  # Assuming yellow_fill is predefined
    ws['Z7'].border = border_negro 
    
    ws['AC7'].alignment = Alignment(horizontal="center", vertical="center")
    ws['AC7'].font = Font(name='Arial', size=8, bold=True, color='000000')
    ws['AC7'].fill = blue_fill  # Assuming yellow_fill is predefined
    ws['AC7'].border = border_negro 
    
    ws['AF7'].alignment = Alignment(horizontal="center", vertical="center")
    ws['AF7'].font = Font(name='Arial', size=8, bold=True, color='000000')
    ws['AF7'].fill = blue_fill  # Assuming yellow_fill is predefined
    ws['AF7'].border = border_negro 
    
    ws['AI7'].alignment = Alignment(horizontal="center", vertical="center")
    ws['AI7'].font = Font(name='Arial', size=8, bold=True, color='000000')
    ws['AI7'].fill = blue_fill  # Assuming yellow_fill is predefined
    ws['AI7'].border = border_negro 
    
    ws['AL7'].alignment = Alignment(horizontal="center", vertical="center")
    ws['AL7'].font = Font(name='Arial', size=8, bold=True, color='000000')
    ws['AL7'].fill = blue_fill  # Assuming yellow_fill is predefined
    ws['AL7'].border = border_negro 
    
    ws['AO7'].alignment = Alignment(horizontal="center", vertical="center")
    ws['AO7'].font = Font(name='Arial', size=8, bold=True, color='000000')
    ws['AO7'].fill = blue_fill  # Assuming yellow_fill is predefined
    ws['AO7'].border = border_negro 
    
    ## crea titulo del reporte
    ws['B1'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B1'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['B1'] = 'OFICINA DE TECNOLOGIAS DE LA INFORMACION'
    
    ws['B2'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B2'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['B2'] = 'DIRECCION REGIONAL DE SALUD JUNIN'
    
    ws['B4'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B4'].font = Font(name = 'Arial', size= 12, bold = True)
    ws['B4'] = 'COBERURA DEL INDICADOR  SI-02.03. PORCENTAJE DE NIÑAS Y NIÑOS DE DOCE (12) MESES DE EDAD, SIN DIAGNÓSTICO DE ANEMIA, QUE RECIBEN TRES (03) DOSAJES DE HEMOGLOBINA Y CULMINAN LA SUPLEMENTACIÓN PREVENTIVA CON HIERRO'
    
    ws['B6'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B6'].font = Font(name = 'Arial', size= 7, bold = True, color='0000CC')
    ws['B6'] ='NOTAS: Primera columna "NUMERADOR", segunda columna "DENOMINADOR", tercera columna "PORCENTAJE AVANCE"'
        
    ws['B8'].alignment = Alignment(horizontal= "center", vertical="center")
    ws['B8'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['B8'].fill = yellow_fill
    ws['B8'].border = border_negro
    ws['B8'] = 'RED'
    
    ws['C8'].alignment = Alignment(horizontal= "center", vertical="center")
    ws['C8'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['C8'].fill = yellow_fill
    ws['C8'].border = border_negro
    ws['C8'] = 'MICRORED'
    
    ws['D8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['D8'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['D8'].fill = yellow_fill
    ws['D8'].border = border_negro
    ws['D8'] = 'ESTABLECIMIENTO'      
    
    ws['E8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['E8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['E8'].fill = yellow_fill
    ws['E8'].border = border_negro
    ws['E8'] = 'N° de niñas y niños del denominador, con suplementación preventiva con hierro y dosaje de hemoglobina, registrados en HIS'
    
    ws['F8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['F8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['F8'].fill = yellow_fill
    ws['F8'].border = border_negro
    ws['F8'] = 'N° de niñas y niños de 12 meses 29 días de edad del departamento, registrados en el padrón nominal con DNI o CNV en línea; y sin diagnóstico de anemia registrado en el HIS a los seis (06) meses de edad'
    
    ws['G8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['G8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['G8'].fill = yellow_fill
    ws['G8'].border = border_negro
    ws['G8'] = '% Avance (Num/Den)'    
    
    ws['H8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['H8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['H8'].fill = blue_fill
    ws['H8'].border = border_negro
    ws['H8'] = 'N° de niñas y niños del denominador, con suplementación preventiva con hierro y dosaje de hemoglobina, registrados en HIS'    
    
    ws['I8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['I8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['I8'].fill = blue_fill
    ws['I8'].border = border_negro
    ws['I8'] = 'N° de niñas y niños de 12 meses 29 días de edad del departamento, registrados en el padrón nominal con DNI o CNV en línea; y sin diagnóstico de anemia registrado en el HIS a los seis (06) meses de edad' 
    
    ws['J8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['J8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['J8'].fill = gray_fill
    ws['J8'].border = border_negro
    ws['J8'] = '% Avance (Num/Den)'    
    
    ws['K8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['K8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['K8'].fill = blue_fill
    ws['K8'].border = border_negro
    ws['K8'] = 'N° de niñas y niños del denominador, con suplementación preventiva con hierro y dosaje de hemoglobina, registrados en HIS'     
    
    ws['L8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['L8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['L8'].fill = blue_fill
    ws['L8'].border = border_negro
    ws['L8'] = 'N° de niñas y niños de 12 meses 29 días de edad del departamento, registrados en el padrón nominal con DNI o CNV en línea; y sin diagnóstico de anemia registrado en el HIS a los seis (06) meses de edad' 
    
    ws['M8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['M8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['M8'].fill = gray_fill
    ws['M8'].border = border_negro
    ws['M8'] = '% Avance (Num/Den)'
    
    ws['N8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['N8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['N8'].fill = blue_fill
    ws['N8'].border = border_negro
    ws['N8'] = 'N° de niñas y niños del denominador, con suplementación preventiva con hierro y dosaje de hemoglobina, registrados en HIS'   
    
    ws['O8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['O8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['O8'].fill = blue_fill
    ws['O8'].border = border_negro
    ws['O8'] = 'N° de niñas y niños de 12 meses 29 días de edad del departamento, registrados en el padrón nominal con DNI o CNV en línea; y sin diagnóstico de anemia registrado en el HIS a los seis (06) meses de edad'
    
    ws['P8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['P8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['P8'].fill = gray_fill
    ws['P8'].border = border_negro
    ws['P8'] = '% Avance (Num/Den)'     
    
    ws['Q8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['Q8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['Q8'].fill = blue_fill
    ws['Q8'].border = border_negro
    ws['Q8'] = 'N° de niñas y niños del denominador, con suplementación preventiva con hierro y dosaje de hemoglobina, registrados en HIS'   
    
    ws['R8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['R8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['R8'].fill = blue_fill
    ws['R8'].border = border_negro
    ws['R8'] = 'N° de niñas y niños de 12 meses 29 días de edad del departamento, registrados en el padrón nominal con DNI o CNV en línea; y sin diagnóstico de anemia registrado en el HIS a los seis (06) meses de edad' 
    
    ws['S8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['S8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['S8'].fill = gray_fill
    ws['S8'].border = border_negro
    ws['S8'] = '% Avance (Num/Den)'    
    
    ws['T8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['T8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['T8'].fill = blue_fill
    ws['T8'].border = border_negro
    ws['T8'] = 'N° de niñas y niños del denominador, con suplementación preventiva con hierro y dosaje de hemoglobina, registrados en HIS'    
    
    ws['U8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['U8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['U8'].fill = blue_fill
    ws['U8'].border = border_negro
    ws['U8'] = 'N° de niñas y niños de 12 meses 29 días de edad del departamento, registrados en el padrón nominal con DNI o CNV en línea; y sin diagnóstico de anemia registrado en el HIS a los seis (06) meses de edad'
    
    ws['V8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['V8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['V8'].fill = gray_fill
    ws['V8'].border = border_negro
    ws['V8'] = '% Avance (Num/Den)'    
    
    ws['W8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['W8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['W8'].fill = blue_fill
    ws['W8'].border = border_negro
    ws['W8'] = 'N° de niñas y niños del denominador, con suplementación preventiva con hierro y dosaje de hemoglobina, registrados en HIS'   
        
    ws['X8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['X8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['X8'].fill = blue_fill
    ws['X8'].border = border_negro
    ws['X8'] = 'N° de niñas y niños de 12 meses 29 días de edad del departamento, registrados en el padrón nominal con DNI o CNV en línea; y sin diagnóstico de anemia registrado en el HIS a los seis (06) meses de edad'

    ws['Y8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['Y8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['Y8'].fill = gray_fill
    ws['Y8'].border = border_negro
    ws['Y8'] = '% Avance (Num/Den)'    
    
    ws['Z8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['Z8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['Z8'].fill = blue_fill
    ws['Z8'].border = border_negro
    ws['Z8'] = 'N° de niñas y niños del denominador, con suplementación preventiva con hierro y dosaje de hemoglobina, registrados en HIS'   

    ws['AA8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AA8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['AA8'].fill = blue_fill
    ws['AA8'].border = border_negro
    ws['AA8'] = 'N° de niñas y niños de 12 meses 29 días de edad del departamento, registrados en el padrón nominal con DNI o CNV en línea; y sin diagnóstico de anemia registrado en el HIS a los seis (06) meses de edad'
    
    ws['AB8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AB8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['AB8'].fill = gray_fill
    ws['AB8'].border = border_negro
    ws['AB8'] = '% Avance (Num/Den)'    
    
    ws['AC8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AC8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['AC8'].fill = blue_fill
    ws['AC8'].border = border_negro
    ws['AC8'] = 'N° de niñas y niños del denominador, con suplementación preventiva con hierro y dosaje de hemoglobina, registrados en HIS'   
    
    ws['AD8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AD8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['AD8'].fill = blue_fill
    ws['AD8'].border = border_negro
    ws['AD8'] = 'N° de niñas y niños de 12 meses 29 días de edad del departamento, registrados en el padrón nominal con DNI o CNV en línea; y sin diagnóstico de anemia registrado en el HIS a los seis (06) meses de edad'
    
    ws['AE8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AE8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['AE8'].fill = gray_fill
    ws['AE8'].border = border_negro
    ws['AE8'] = '% Avance (Num/Den)'    
    
    ws['AF8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AF8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['AF8'].fill = blue_fill
    ws['AF8'].border = border_negro
    ws['AF8'] = 'N° de niñas y niños del denominador, con suplementación preventiva con hierro y dosaje de hemoglobina, registrados en HIS'   
    
    ws['AG8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AG8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['AG8'].fill = blue_fill
    ws['AG8'].border = border_negro
    ws['AG8'] = 'N° de niñas y niños de 12 meses 29 días de edad del departamento, registrados en el padrón nominal con DNI o CNV en línea; y sin diagnóstico de anemia registrado en el HIS a los seis (06) meses de edad' 
    
    ws['AH8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AH8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['AH8'].fill = gray_fill
    ws['AH8'].border = border_negro
    ws['AH8'] = '% Avance (Num/Den)'    
    
    ws['AI8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AI8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['AI8'].fill = blue_fill
    ws['AI8'].border = border_negro
    ws['AI8'] = 'N° de niñas y niños del denominador, con suplementación preventiva con hierro y dosaje de hemoglobina, registrados en HIS'    
    
    ws['AJ8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AJ8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['AJ8'].fill = blue_fill
    ws['AJ8'].border = border_negro
    ws['AJ8'] = 'N° de niñas y niños de 12 meses 29 días de edad del departamento, registrados en el padrón nominal con DNI o CNV en línea; y sin diagnóstico de anemia registrado en el HIS a los seis (06) meses de edad'
    
    ws['AK8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AK8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['AK8'].fill = gray_fill
    ws['AK8'].border = border_negro
    ws['AK8'] = '% Avance (Num/Den)'    
    
    ws['AL8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AL8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['AL8'].fill = blue_fill
    ws['AL8'].border = border_negro
    ws['AL8'] = 'N° de niñas y niños del denominador, con suplementación preventiva con hierro y dosaje de hemoglobina, registrados en HIS'    
    
    ws['AM8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AM8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['AM8'].fill = blue_fill
    ws['AM8'].border = border_negro
    ws['AM8'] = 'N° de niñas y niños de 12 meses 29 días de edad del departamento, registrados en el padrón nominal con DNI o CNV en línea; y sin diagnóstico de anemia registrado en el HIS a los seis (06) meses de edad'
    
    ws['AN8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AN8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['AN8'].fill = gray_fill
    ws['AN8'].border = border_negro
    ws['AN8'] = '% Avance (Num/Den)'    
    
    ws['AO8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AO8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['AO8'].fill = blue_fill
    ws['AO8'].border = border_negro
    ws['AO8'] = 'N° de niñas y niños del denominador, con suplementación preventiva con hierro y dosaje de hemoglobina, registrados en HIS'   
    
    ws['AP8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AP8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['AP8'].fill = blue_fill
    ws['AP8'].border = border_negro
    ws['AP8'] = 'N° de niñas y niños de 12 meses 29 días de edad del departamento, registrados en el padrón nominal con DNI o CNV en línea; y sin diagnóstico de anemia registrado en el HIS a los seis (06) meses de edad' 
    
    ws['AQ8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AQ8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['AQ8'].fill = gray_fill
    ws['AQ8'].border = border_negro
    ws['AQ8'] = '% Avance (Num/Den)'    
    
    # Definir estilos
    header_font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    centered_alignment = Alignment(horizontal='center')
    border = Border(left=Side(style='thin', color='A9A9A9'),
            right=Side(style='thin', color='A9A9A9'),
            top=Side(style='thin', color='A9A9A9'),
            bottom=Side(style='thin', color='A9A9A9'))
    header_fill = PatternFill(patternType='solid', fgColor='00B0F0')
    
    # Definir los caracteres especiales de check y X
    
    # Escribir datos
    for row, record in enumerate(results, start=9):
        for col, value in enumerate(record, start=2):
            cell = ws.cell(row=row, column=col, value=value)

            # Alinear a la izquierda solo en las columnas 6,14,15,16
            if col in [2, 3, 4]:
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='center')

            # Aplicar formato basado en el valor para columnas específicas
            if col in [7, 10, 13, 16, 19, 22, 25, 28, 31, 34, 37, 40, 43]:
                try:
                    value_float = float(value)
                except (ValueError, TypeError):
                    # Si el valor no es numérico, aplicar formato por defecto
                    cell.font = Font(name='Arial', size=7)
                else:
                    # Si los valores están entre 0 y 100, dividimos entre 100
                    if value_float > 1:
                        value_float = value_float / 100
                        cell.value = value_float

                    # Establecer el formato de número a porcentaje
                    cell.number_format = '0.0%'

                    if value_float >= 0.80:
                        # Colorear la celda de verde
                        cell.fill = PatternFill(patternType='solid', fgColor='00B050')  # Fondo verde
                        cell.font = Font(name='Arial', size=7, color='000000')  # Letra negra
                    else:
                        # Colorear la celda de rojo con letras blancas
                        cell.fill = PatternFill(patternType='solid', fgColor='FF0000')  # Fondo rojo
                        cell.font = Font(name='Arial', size=7, color='FFFFFF')  # Letra blanca
            # Fuente normal para otras columnas
            else:
                cell.font = Font(name='Arial', size=8)
            
            cell.border = border