from django.shortcuts import render

# TABLERO Adolescente Hemopglobina 
from django.db import connection
from django.http import JsonResponse
from base.models import MAESTRO_HIS_ESTABLECIMIENTO, DimPeriodo
from django.db.models.functions import Substr
import logging

# report excel
from django.http.response import HttpResponse
from django.views.generic.base import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import openpyxl
from openpyxl.utils import get_column_letter

from django.db.models.functions import Substr

from datetime import datetime
import locale

from django.db.models import IntegerField               # Importar IntegerField
from django.db.models.functions import Cast, Substr     # Importar Cast y Substr

logger = logging.getLogger(__name__)

# Reporte excel
from datetime import datetime
import getpass  # Para obtener el nombre del usuario
from django.contrib.auth.models import User  # O tu modelo de usuario personalizado
from django.http import HttpResponse
from io import BytesIO
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required

User = get_user_model()

from django.db.models import IntegerField               # Importar IntegerField
from django.db.models.functions import Cast, Substr     # Importar Cast y Substr
from base.models import Actualizacion
# Create your views here.
def obtener_distritos(provincia):
    distritos = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Provincia=provincia).values('Distrito').distinct().order_by('Distrito')
    return list(distritos)

def obtener_avance_v1_condicion_previa(red):
    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT * FROM public.obtener_avance_v1_condicion_previa(%s)",
            [red]
        )
        return cursor.fetchall()

def obtener_ranking_v1_condicion_previa(anio, mes):
    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT * FROM public.obtener_ranking_v1_condicion_previa(%s, %s)",
            [anio, mes]
        )
        result = cursor.fetchall()
        return result

## AVANCE REGIONAL
def obtener_avance_regional_v1_condicion_previa():
    """
    Obtiene el avance regional de gestantes con anemia.
    Retorna una lista de diccionarios con las claves 'num', 'den' y 'cob'.
    """
    try:
        # Asegúrate de que la conexión a la base de datos está establecida
        with connection.cursor() as cursor:
            cursor.execute(
                '''
                SELECT 
                    SUM(numerador_mes_diciembre) AS num,
                    SUM(denominador_mes_diciembre) AS den,
                    ROUND((SUM(numerador_mes_diciembre)::NUMERIC / NULLIF(SUM(denominador_mes_diciembre), 0)) * 100, 2) AS cob
                FROM public."Cobertura_VI0101_CondicionPrevia"
                '''
            )
            resultados = cursor.fetchall()
            
            # Obtener los nombres de las columnas
            column_names = [desc[0] for desc in cursor.description]
            
            # Convertir cada fila en un diccionario
            datos = [dict(zip(column_names, fila)) for fila in resultados]
        
        return datos
    except Exception as e:
        print(f"Error al obtener el avance regional: {e}")
        return None

## AVANCE REGIONAL MENSUALIZADO
def obtener_avance_regional_mensual_v1_condicion_previa(anio):
    """
    Obtiene el avance regional de gestantes con anemia de manera mensualizada.
    Retorna una lista de diccionarios con las claves 'num', 'den' y 'cob' por meses.
    """
    try:
        # Asegúrate de que la conexión a la base de datos está establecida
        with connection.cursor() as cursor:
            cursor.execute(
                '''
                                        SELECT
                    -- ENERO
                    SUM(CASE WHEN mes = 1 THEN CAST(numerador AS INT) ELSE 0 END) AS NUM_1,
                    SUM(CASE WHEN mes = 1 THEN CAST(denominador AS INT) ELSE 0 END) AS DEN_1,
                    CASE 
                        WHEN SUM(CASE WHEN mes = 1 THEN CAST(denominador AS INT) ELSE 0 END) = 0 
                        THEN 0 
                        ELSE ROUND(
                            (
                                SUM(CASE WHEN mes = 1 THEN CAST(numerador AS INT) ELSE 0 END) * 1.0 
                                / NULLIF(SUM(CASE WHEN mes = 1 THEN CAST(denominador AS INT) ELSE 0 END), 0)
                            ) * 100
                        , 2) 
                    END AS COB_1,
                    -- FEBRERO
                    SUM(CASE WHEN mes = 2 THEN CAST(numerador AS INT) ELSE 0 END) AS NUM_2,
                    SUM(CASE WHEN mes = 2 THEN CAST(denominador AS INT) ELSE 0 END) AS DEN_2,
                    CASE 
                        WHEN SUM(CASE WHEN mes = 2 THEN CAST(denominador AS INT) ELSE 0 END) = 0 
                        THEN 0 
                        ELSE ROUND(
                            (
                                SUM(CASE WHEN mes = 2 THEN CAST(numerador AS INT) ELSE 0 END) * 1.0 
                                / NULLIF(SUM(CASE WHEN mes = 2 THEN CAST(denominador AS INT) ELSE 0 END), 0)
                            ) * 100
                        , 2) 
                    END AS COB_2,
                    -- MARZO
                    SUM(CASE WHEN mes = 3 THEN CAST(numerador AS INT) ELSE 0 END) AS NUM_3,
                    SUM(CASE WHEN mes = 3 THEN CAST(denominador AS INT) ELSE 0 END) AS DEN_3,
                    CASE 
                        WHEN SUM(CASE WHEN mes = 3 THEN CAST(denominador AS INT) ELSE 0 END) = 0 
                        THEN 0 
                        ELSE ROUND(
                            (
                                SUM(CASE WHEN mes = 3 THEN CAST(numerador AS INT) ELSE 0 END) * 1.0 
                                / NULLIF(SUM(CASE WHEN mes = 3 THEN CAST(denominador AS INT) ELSE 0 END), 0)
                            ) * 100
                        , 2) 
                    END AS COB_3,
                    -- ABRIL
                    SUM(CASE WHEN mes = 4 THEN CAST(numerador AS INT) ELSE 0 END) AS NUM_4,
                    SUM(CASE WHEN mes = 4 THEN CAST(denominador AS INT) ELSE 0 END) AS DEN_4,
                    CASE 
                        WHEN SUM(CASE WHEN mes = 4 THEN CAST(denominador AS INT) ELSE 0 END) = 0 
                        THEN 0 
                        ELSE ROUND(
                            (
                                SUM(CASE WHEN mes = 4 THEN CAST(numerador AS INT) ELSE 0 END) * 1.0 
                                / NULLIF(SUM(CASE WHEN mes = 4 THEN CAST(denominador AS INT) ELSE 0 END), 0)
                            ) * 100
                        , 2) 
                    END AS COB_4,
                    -- MAYO
                    SUM(CASE WHEN mes = 5 THEN CAST(numerador AS INT) ELSE 0 END) AS NUM_5,
                    SUM(CASE WHEN mes = 5 THEN CAST(denominador AS INT) ELSE 0 END) AS DEN_5,
                    CASE 
                        WHEN SUM(CASE WHEN mes = 5 THEN CAST(denominador AS INT) ELSE 0 END) = 0 
                        THEN 0 
                        ELSE ROUND(
                            (
                                SUM(CASE WHEN mes = 5 THEN CAST(numerador AS INT) ELSE 0 END) * 1.0
                                / NULLIF(SUM(CASE WHEN mes = 5 THEN CAST(denominador AS INT) ELSE 0 END), 0)
                            ) * 100
                        , 2) 
                    END AS COB_5,
                    -- JUNIO
                    SUM(CASE WHEN mes = 6 THEN CAST(numerador AS INT) ELSE 0 END) AS NUM_6,
                    SUM(CASE WHEN mes = 6 THEN CAST(denominador AS INT) ELSE 0 END) AS DEN_6,
                    CASE 
                        WHEN SUM(CASE WHEN mes = 6 THEN CAST(denominador AS INT) ELSE 0 END) = 0 
                        THEN 0 
                        ELSE ROUND(
                            (
                                SUM(CASE WHEN mes = 6 THEN CAST(numerador AS INT) ELSE 0 END) * 1.0 
                                / NULLIF(SUM(CASE WHEN mes = 6 THEN CAST(denominador AS INT) ELSE 0 END), 0)
                            ) * 100
                        , 2) 
                    END AS COB_6,
                    -- JULIO
                    SUM(CASE WHEN mes = 7 THEN CAST(numerador AS INT) ELSE 0 END) AS NUM_7,
                    SUM(CASE WHEN mes = 7 THEN CAST(denominador AS INT) ELSE 0 END) AS DEN_7,
                    CASE 
                        WHEN SUM(CASE WHEN mes = 7 THEN CAST(denominador AS INT) ELSE 0 END) = 0 
                        THEN 0 
                        ELSE ROUND(
                            (
                                SUM(CASE WHEN mes = 7 THEN CAST(numerador AS INT) ELSE 0 END) * 1.0 
                                / NULLIF(SUM(CASE WHEN mes = 7 THEN CAST(denominador AS INT) ELSE 0 END), 0)
                            ) * 100
                        , 2) 
                    END AS COB_7,
                    -- AGOSTO
                    SUM(CASE WHEN mes = 8 THEN CAST(numerador AS INT) ELSE 0 END) AS NUM_8,
                    SUM(CASE WHEN mes = 8 THEN CAST(denominador AS INT) ELSE 0 END) AS DEN_8,
                    CASE 
                        WHEN SUM(CASE WHEN mes = 8 THEN CAST(denominador AS INT) ELSE 0 END) = 0 
                        THEN 0 
                        ELSE ROUND(
                            (
                                SUM(CASE WHEN mes = 8 THEN CAST(numerador AS INT) ELSE 0 END) * 1.0 
                                / NULLIF(SUM(CASE WHEN mes = 8 THEN CAST(denominador AS INT) ELSE 0 END), 0)
                            ) * 100
                        , 2) 
                    END AS COB_8,
                    -- SETIEMBRE
                    SUM(CASE WHEN mes = 9 THEN CAST(numerador AS INT) ELSE 0 END) AS NUM_9,
                    SUM(CASE WHEN mes = 9 THEN CAST(denominador AS INT) ELSE 0 END) AS DEN_9,
                    CASE 
                        WHEN SUM(CASE WHEN mes = 9 THEN CAST(denominador AS INT) ELSE 0 END) = 0 
                        THEN 0 
                        ELSE ROUND(
                            (
                                SUM(CASE WHEN mes = 9 THEN CAST(numerador AS INT) ELSE 0 END) * 1.0 
                                / NULLIF(SUM(CASE WHEN mes = 9 THEN CAST(denominador AS INT) ELSE 0 END), 0)
                            ) * 100
                        , 2)
                    END AS COB_9,
                    -- OCTUBRE
                    SUM(CASE WHEN mes = 10 THEN CAST(numerador AS INT) ELSE 0 END) AS NUM_10,
                    SUM(CASE WHEN mes = 10 THEN CAST(denominador AS INT) ELSE 0 END) AS DEN_10,
                    CASE 
                        WHEN SUM(CASE WHEN mes = 10 THEN CAST(denominador AS INT) ELSE 0 END) = 0 
                        THEN 0 
                        ELSE ROUND(
                            (
                                SUM(CASE WHEN mes = 10 THEN CAST(numerador AS INT) ELSE 0 END) * 1.0 
                                / NULLIF(SUM(CASE WHEN mes = 10 THEN CAST(denominador AS INT) ELSE 0 END), 0)
                            ) * 100
                        , 2) 
                    END AS COB_10,
                    -- NOVIEMBRE
                    SUM(CASE WHEN mes = 11 THEN CAST(numerador AS INT) ELSE 0 END) AS NUM_11,
                    SUM(CASE WHEN mes = 11 THEN CAST(denominador AS INT) ELSE 0 END) AS DEN_11,
                    CASE 
                        WHEN SUM(CASE WHEN mes = 11 THEN CAST(denominador AS INT) ELSE 0 END) = 0 
                        THEN 0 
                        ELSE ROUND(
                            (
                                SUM(CASE WHEN mes = 11 THEN CAST(numerador AS INT) ELSE 0 END) * 1.0 
                                / NULLIF(SUM(CASE WHEN mes = 11 THEN CAST(denominador AS INT) ELSE 0 END), 0)
                            ) * 100
                        , 2) 
                    END AS COB_11,
                    -- DICIEMBRE
                    SUM(CASE WHEN mes = 12 THEN CAST(numerador AS INT) ELSE 0 END) AS NUM_12,
                    SUM(CASE WHEN mes = 12 THEN CAST(denominador AS INT) ELSE 0 END) AS DEN_12,
                    CASE 
                        WHEN SUM(CASE WHEN mes = 12 THEN CAST(denominador AS INT) ELSE 0 END) = 0 
                        THEN 0 
                        ELSE ROUND(
                            (
                                SUM(CASE WHEN mes = 12 THEN CAST(numerador AS INT) ELSE 0 END) * 1.0 
                                / NULLIF(SUM(CASE WHEN mes = 12 THEN CAST(denominador AS INT) ELSE 0 END), 0)
                            ) * 100
                        , 2) 
                    END AS COB_12
                    FROM public."VI0101_CondicionPrevia_Combinado"
                    WHERE año = %s;
                ''',
                [anio]  # Filtro por año
            )
            resultados = cursor.fetchall()
            
            # Obtener los nombres de las columnas
            column_names = [desc[0] for desc in cursor.description]
            
            # Convertir cada fila en un diccionario
            datos = [dict(zip(column_names, fila)) for fila in resultados]
        
        return datos
    except Exception as e:
        print(f"Error al obtener el avance regional: {e}")
        return None

def index_v1_condicion_previa(request):
    actualizacion = Actualizacion.objects.all()
    # RANKING 
        # Capturamos el año que viene por GET
    anio = request.GET.get('anio', None)
    if anio not in ['2024', '2025']:
        # Si no llega un año válido, puedes fijar uno por defecto (2024, por ejemplo)
        anio = '2024'
    
    mes_seleccionado = request.GET.get('mes')
    # GRAFICO
    red_seleccionada = request.GET.get('red')
    red = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Disa='JUNIN').values_list('Red', flat=True).distinct().order_by('Red')
    # Si la solicitud es AJAX
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            # Obtener datos de RANKING 
            resultados_ranking_obtener_v1_condicion_previa = obtener_ranking_v1_condicion_previa(anio,mes_seleccionado)
            # Obtener datos de AVANCE GRAFICO MESES
            resultados_avance_obtener_v1_condicion_previa = obtener_avance_v1_condicion_previa(red_seleccionada)
            # Obtener datos de AVANCE GRAFICO ANUAL
            resultados_avance_regional_v1_condicion_previa = obtener_avance_regional_v1_condicion_previa()
            # Obtener datos de AVANCE GRAFICO ANUAL
            resultados_avance_regional_mensual_v1_condicion_previa = obtener_avance_regional_mensual_v1_condicion_previa(anio)
            
            # Procesar los resultados
            if any(len(row) < 4 for row in resultados_ranking_obtener_v1_condicion_previa):
                raise ValueError("Algunas filas del ranking no tienen suficientes elementos")
            
            data = {               
                #ranking
                'red': [],
                'num_r': [],
                'den_r': [],
                'avance_r': [],
                
                #avance meses
                'mes': [],
                'num': [],
                'den': [],
                'avance': [],

                'num_region': [],
                'den_region': [],
                'avance_region': [],
                
                #avance regional mensual
                'num_1': [],
                'den_1': [],
                'cob_1': [],
                'num_2': [],
                'den_2': [],
                'cob_2': [],
                'num_3': [],
                'den_3': [],
                'cob_3': [],
                'num_4': [],
                'den_4': [],
                'cob_4': [],
                'num_5': [],
                'den_5': [],
                'cob_5': [],
                'num_6': [],
                'den_6': [],
                'cob_6': [],
                'num_7': [],
                'den_7': [],
                'cob_7': [],
                'num_8': [],
                'den_8': [],
                'cob_8': [],                
                'num_9': [],
                'den_9': [],
                'cob_9': [],
                'num_10': [],
                'den_10': [],
                'cob_10': [],
                'num_11': [],
                'den_11': [],
                'cob_11': [],
                'num_12': [],
                'den_12': [],
                'cob_12': [],
            }     
            # AVANCE GRAFICO REGIONAL
            for index, row in enumerate(resultados_avance_regional_v1_condicion_previa):
                try:
                    # Verifica que el diccionario tenga las claves necesarias
                    required_keys = {'num', 'den', 'cob'}
                    if not required_keys.issubset(row.keys()):
                        raise ValueError(f"La fila {index} no tiene las claves necesarias: {row}")

                    num_region_value = float(row.get('num', 0.0))
                    den_region_value = float(row.get('den', 0.0))
                    avance_region_value = float(row.get('cob', 0.0))

                    data['num_region'].append(num_region_value)
                    data['den_region'].append(den_region_value)
                    data['avance_region'].append(avance_region_value)

                except Exception as e:
                    logger.error(f"Error procesando la fila {index}: {str(e)}")
                    
            # AVANCE GRAFICO MENSUAL
            for index, row in enumerate(resultados_avance_regional_mensual_v1_condicion_previa):
                try:
                    # Verifica que el diccionario tenga las claves necesarias
                    required_keys = {'num_1','den_1','cob_1','num_2','den_2','cob_2','num_3','den_3','cob_3','num_4','den_4','cob_4','num_5','den_5','cob_5','num_6','den_6','cob_6','num_7','den_7','cob_7','num_8','den_8','cob_8','num_9','den_9','cob_9','num_10','den_10','cob_10','num_11','den_11','cob_11','num_12','den_12','cob_12'}
                    
                    if not required_keys.issubset(row.keys()):
                        raise ValueError(f"La fila {index} no tiene las claves necesarias: {row}")

                    num_1_value = float(row.get('num_1', 0.0))
                    den_1_value = float(row.get('den_1', 0.0))
                    cob_1_value = float(row.get('cob_1', 0.0))
                    num_2_value = float(row.get('num_2', 0.0))
                    den_2_value = float(row.get('den_2', 0.0))
                    cob_2_value = float(row.get('cob_2', 0.0))
                    num_3_value = float(row.get('num_3', 0.0))
                    den_3_value = float(row.get('den_3', 0.0))
                    cob_3_value = float(row.get('cob_3', 0.0))
                    num_4_value = float(row.get('num_4', 0.0))
                    den_4_value = float(row.get('den_4', 0.0))
                    cob_4_value = float(row.get('cob_4', 0.0))
                    num_5_value = float(row.get('num_5', 0.0))
                    den_5_value = float(row.get('den_5', 0.0))
                    cob_5_value = float(row.get('cob_5', 0.0))
                    num_6_value = float(row.get('num_6', 0.0))
                    den_6_value = float(row.get('den_6', 0.0))
                    cob_6_value = float(row.get('cob_6', 0.0))
                    num_7_value = float(row.get('num_7', 0.0))
                    den_7_value = float(row.get('den_7', 0.0))
                    cob_7_value = float(row.get('cob_7', 0.0))
                    num_8_value = float(row.get('num_8', 0.0))
                    den_8_value = float(row.get('den_8', 0.0))
                    cob_8_value = float(row.get('cob_8', 0.0))
                    num_9_value = float(row.get('num_9', 0.0))
                    den_9_value = float(row.get('den_9', 0.0))
                    cob_9_value = float(row.get('cob_9', 0.0))
                    num_10_value = float(row.get('num_10', 0.0))
                    den_10_value = float(row.get('den_10', 0.0))
                    cob_10_value = float(row.get('cob_10', 0.0))
                    num_11_value = float(row.get('num_11', 0.0))
                    den_11_value = float(row.get('den_11', 0.0))
                    cob_11_value = float(row.get('cob_11', 0.0))
                    num_12_value = float(row.get('num_12', 0.0))
                    den_12_value = float(row.get('den_12', 0.0))
                    cob_12_value = float(row.get('cob_12', 0.0))
                    
                    data['num_1'].append(num_1_value)
                    data['den_1'].append(den_1_value)
                    data['cob_1'].append(cob_1_value)
                    data['num_2'].append(num_2_value)
                    data['den_2'].append(den_2_value)
                    data['cob_2'].append(cob_2_value)
                    data['num_3'].append(num_3_value)
                    data['den_3'].append(den_3_value)
                    data['cob_3'].append(cob_3_value)
                    data['num_4'].append(num_4_value)
                    data['den_4'].append(den_4_value)
                    data['cob_4'].append(cob_4_value)
                    data['num_5'].append(num_5_value)
                    data['den_5'].append(den_5_value)
                    data['cob_5'].append(cob_5_value)
                    data['num_6'].append(num_6_value)
                    data['den_6'].append(den_6_value)
                    data['cob_6'].append(cob_6_value)
                    data['num_7'].append(num_7_value)
                    data['den_7'].append(den_7_value)
                    data['cob_7'].append(cob_7_value)
                    data['num_8'].append(num_8_value)
                    data['den_8'].append(den_8_value)
                    data['cob_8'].append(cob_8_value)
                    data['num_9'].append(num_9_value)
                    data['den_9'].append(den_9_value)
                    data['cob_9'].append(cob_9_value)
                    data['num_10'].append(num_10_value)
                    data['den_10'].append(den_10_value)
                    data['cob_10'].append(cob_10_value)
                    data['num_11'].append(num_11_value)
                    data['den_11'].append(den_11_value)
                    data['cob_11'].append(cob_11_value)
                    data['num_12'].append(num_12_value)
                    data['den_12'].append(den_12_value)
                    data['cob_12'].append(cob_12_value)

                except Exception as e:
                    logger.error(f"Error procesando la fila {index}: {str(e)}")
            #RANKING
            for index, row in enumerate(resultados_ranking_obtener_v1_condicion_previa):
                try:
                    # Verifica que la tupla tenga exactamente 4 elementos
                    if len(row) != 4:
                        raise ValueError(f"La fila {index} no tiene 4 elementos: {row}")

                    red_value = row[0] if row[0] is not None else ''
                    num_r_value = float(row[1]) if row[1] is not None else 0.0
                    den_r_value = float(row[2]) if row[2] is not None else 0.0
                    avance_r_value = float(row[3]) if row[3] is not None else 0.0

                    data['red'].append(red_value)
                    data['num_r'].append(num_r_value)
                    data['den_r'].append(den_r_value)
                    data['avance_r'].append(avance_r_value)

                except Exception as e:
                    logger.error(f"Error procesando la fila {index}: {str(e)}")
            
            #AVANCE GRAFICO MESES
            for index, row in enumerate(resultados_avance_obtener_v1_condicion_previa):
                try:
                    # Verifica que la tupla tenga exactamente 4 elementos
                    if len(row) != 5:
                        raise ValueError(f"La fila {index} no tiene 5 elementos: {row}")

                    mes_value = row[1] if row[1] is not None else ''
                    num_value = float(row[2]) if row[2] is not None else 0.0
                    den_value = float(row[3]) if row[3] is not None else 0.0
                    avance_value = float(row[4]) if row[4] is not None else 0.0

                    data['mes'].append(mes_value)
                    data['num'].append(num_value)
                    data['den'].append(den_value)
                    data['avance'].append(avance_value)

                except Exception as e:
                    logger.error(f"Error procesando la fila {index}: {str(e)}")
            
            return JsonResponse(data)

        except Exception as e:
            logger.error(f"Error al obtener datos: {str(e)}")

    # Si no es una solicitud AJAX, renderiza la página principal
    return render(request, 'v1_condicion_previa/index_v1_condicion_previa.html', {
        'red': red,
        'mes_seleccionado': mes_seleccionado,
        'actualizacion': actualizacion
    })

## SEGUIMIENTO
def get_redes_v1_condicion_previa(request,redes_id):
    redes = (
            MAESTRO_HIS_ESTABLECIMIENTO
            .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL',Departamento='JUNIN')
            .annotate(codigo_red_filtrado=Substr('Codigo_Red', 1, 4))
            .values('Red','codigo_red_filtrado')
            .distinct()
            .order_by('Red')
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'redes': redes,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
    }
    
    return render(request, 'v1_condicion_previa/redes.html', context)

## SEGUIMIENTO POR MICRO-REDES
def get_microredes_v1_condicion_previa(request, microredes_id):
    redes = (
            MAESTRO_HIS_ESTABLECIMIENTO
            .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL',Departamento='JUNIN')
            .annotate(codigo_red_filtrado=Substr('Codigo_Red', 1, 4))
            .values('Red','codigo_red_filtrado')
            .distinct()
            .order_by('Red')
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'redes': redes,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
    }
    
    return render(request, 'v1_condicion_previa/microredes.html', context)

def p_microredes_v1_condicion_previa(request):
    redes_param = request.GET.get('red')
    microredes = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Codigo_Red=redes_param, Descripcion_Sector='GOBIERNO REGIONAL', Disa='JUNIN').values('Codigo_MicroRed','MicroRed').distinct()
    context = {
        'redes_param': redes_param,
        'microredes': microredes
    }
    return render(request, 'v1_condicion_previa/partials/p_microredes.html', context)

## REPORTE POR ESTABLECIMIENTO
def get_establecimientos_v1_condicion_previa(request,establecimiento_id):
    redes = (
                MAESTRO_HIS_ESTABLECIMIENTO
                .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL',Disa='JUNIN')
                .annotate(codigo_red_filtrado=Substr('Codigo_Red', 1, 4))
                .values('Red','codigo_red_filtrado')
                .distinct()
                .order_by('Red')
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'redes': redes,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
    }
    return render(request,'v1_condicion_previa/establecimientos.html', context)

def p_microredes_establec_v1_condicion_previa(request):
    redes_param = request.GET.get('red') 
    microredes = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Codigo_Red=redes_param, Descripcion_Sector='GOBIERNO REGIONAL',Disa='JUNIN').values('Codigo_MicroRed','MicroRed').distinct()
    context = {
        'microredes': microredes,
        'is_htmx': True
    }
    return render(request, 'v1_condicion_previa/partials/p_microredes_establec.html', context)

def p_establecimientos_v1_condicion_previa(request):
    microredes = request.GET.get('p_microredes_establec')    
    codigo_red = request.GET.get('red')
    establec = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Codigo_MicroRed=microredes,Codigo_Red=codigo_red,Descripcion_Sector='GOBIERNO REGIONAL',Disa='JUNIN').values('Codigo_Unico','Nombre_Establecimiento').distinct()

    context= {
        'establec': establec
    }
    return render(request, 'v1_condicion_previa/partials/p_establecimientos.html', context)

## QUERY
def obtener_seguimiento_redes_v1_condicion_previa(p_anio,p_red,p_microred,p_establec,p_inicio,p_fin,p_cumple):
    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT * FROM public.fn_seguimiento_v1_condicion_previa(%s,%s,%s,%s,%s,%s,%s)",
            [p_anio, p_red, p_microred, p_establec, p_inicio, p_fin, p_cumple]
        )
        return cursor.fetchall()

## TEMPLATES 
class RptV1CondicionPreviaRed(TemplateView):
    def get(self, request, *args, **kwargs):
        # Variables ingresadas
        p_anio = request.GET.get('anio')
        p_red = request.GET.get('red','')
        p_microred = ''
        p_establec = ''
        p_inicio = int(request.GET.get('fecha_inicio'))
        p_fin = int(request.GET.get('fecha_fin'))
        p_cumple = request.GET.get('cumple', '')     
        # Creación de la consulta
        resultado_seguimiento = obtener_seguimiento_redes_v1_condicion_previa(p_anio,p_red,p_microred,p_establec,p_inicio,p_fin,p_cumple)
        
        wb = Workbook()
        
        consultas = [
                ('Seguimiento', resultado_seguimiento)
        ]
        
        for index, (sheet_name, results) in enumerate(consultas):
            if index == 0:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
        
            fill_worksheet_v1_condicion_previa(ws, results)
        ##########################################################################          
        # Establecer el nombre del archivo
        nombre_archivo = "rpt_v1_condicion_previa.xlsx"
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)

        return response

class RptV1CondicionPreviaMicroRed(TemplateView):
    def get(self, request, *args, **kwargs):
        # Variables ingresadas
        p_anio = request.GET.get('anio')
        p_red = request.GET.get('red','')
        p_microred = request.GET.get('p_microredes','')
        p_establec = ''
        p_inicio = int(request.GET.get('fecha_inicio'))
        p_fin = int(request.GET.get('fecha_fin'))
        p_cumple = request.GET.get('cumple', '')     
        # Creación de la consulta
        resultado_seguimiento = obtener_seguimiento_redes_v1_condicion_previa(p_anio,p_red,p_microred,p_establec,p_inicio,p_fin,p_cumple)
                
        wb = Workbook()
        
        consultas = [
                ('Seguimiento', resultado_seguimiento)
        ]
        
        for index, (sheet_name, results) in enumerate(consultas):
            if index == 0:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
        
            fill_worksheet_v1_condicion_previa(ws, results)
        
        ##########################################################################          
        # Establecer el nombre del archivo
        nombre_archivo = "rpt_v1_condicion_previa_red.xlsx"
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)

        return response

class RptV1CondicionPreviaEstablec(TemplateView):
    def get(self, request, *args, **kwargs):
        # Variables ingresadas
        p_anio = request.GET.get('anio')
        p_red = request.GET.get('red','')
        p_microred = request.GET.get('p_microredes','')
        p_establec = request.GET.get('p_establecimiento','')
        p_inicio = int(request.GET.get('fecha_inicio'))
        p_fin = int(request.GET.get('fecha_fin'))
        p_cumple = request.GET.get('cumple', '')     
        # Creación de la consulta
        resultado_seguimiento = obtener_seguimiento_redes_v1_condicion_previa(p_anio,p_red,p_microred,p_establec,p_inicio,p_fin,p_cumple)
                
        wb = Workbook()
        
        consultas = [
                ('Seguimiento', resultado_seguimiento)
        ]
        
        for index, (sheet_name, results) in enumerate(consultas):
            if index == 0:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
        
            fill_worksheet_v1_condicion_previa(ws, results)
        
        ##########################################################################          
        # Establecer el nombre del archivo
        nombre_archivo = "rpt_v1_condicion_previa_red.xlsx"
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)

        return response

def fill_worksheet_v1_condicion_previa(ws, results): 
    
    ws.row_dimensions[1].height = 14
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 12
    ws.row_dimensions[4].height = 25
    ws.row_dimensions[5].height = 20
    ws.row_dimensions[6].height = 60
    ws.row_dimensions[7].height = 60
    # cambia el ancho de la columna
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 9
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 9
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 9
    ws.column_dimensions['I'].width = 16    
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 16
    ws.column_dimensions['L'].width = 20
    ws.column_dimensions['M'].width = 9
    ws.column_dimensions['N'].width = 30

    # linea de division
    ws.freeze_panes = 'D9'
    # Configuración del fondo y el borde
    # Definir el color usando formato aRGB (opacidad completa 'FF' + color RGB)
    fill = PatternFill(start_color='FF60D7E0', end_color='FF60D7E0', fill_type='solid')
    # Definir el color anaranjado usando formato aRGB
    orange_fill = PatternFill(start_color='FFE0A960', end_color='FFE0A960', fill_type='solid')
    # Definir los estilos para gris
    gray_fill = PatternFill(start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid')
    # Definir el estilo de color verde
    green_fill = PatternFill(start_color='FF60E0B3', end_color='FF60E0B3', fill_type='solid')
    # Definir el estilo de color amarillo
    yellow_fill = PatternFill(start_color='FFE0DE60', end_color='FFE0DE60', fill_type='solid')
    # Definir el estilo de color azul
    blue_fill = PatternFill(start_color='FF60A2E0', end_color='FF60A2E0', fill_type='solid')
    # Definir el estilo de color verde 2
    green_fill_2 = PatternFill(start_color='FF60E07E', end_color='FF60E07E', fill_type='solid')   
    
    green_font = Font(name='Arial', size=8, color='00FF00')  # Verde
    red_font = Font(name='Arial', size=8, color='FF0000')    # Rojo
    
    border = Border(left=Side(style='thin', color='00B0F0'),
                    right=Side(style='thin', color='00B0F0'),
                    top=Side(style='thin', color='00B0F0'),
                    bottom=Side(style='thin', color='00B0F0'))
    borde_plomo = Border(left=Side(style='thin', color='A9A9A9'), # Plomo
                    right=Side(style='thin', color='A9A9A9'), # Plomo
                    top=Side(style='thin', color='A9A9A9'), # Plomo
                    bottom=Side(style='thin', color='A9A9A9')) # Plomo
    
    # Definir el estilo de relleno celeste
    celeste_fill = PatternFill(start_color='FF87CEEB', end_color='FF87CEEB', fill_type='solid')
    # Morado más claro
    morado_claro_fill = PatternFill(start_color='FFE9D8FF', end_color='FFE9D8FF', fill_type='solid')
    # Plomo más claro
    plomo_claro_fill = PatternFill(start_color='FFEDEDED', end_color='FFEDEDED', fill_type='solid')
    # Azul más claro
    azul_claro_fill = PatternFill(start_color='FFD8EFFA', end_color='FFD8EFFA', fill_type='solid')
    # Naranja más claro
    naranja_claro_fill = PatternFill(start_color='FFFFEBD8', end_color='FFFFEBD8', fill_type='solid')
    # Verde más claro
    verde_claro_fill = PatternFill(start_color='FFBDF7BD', end_color='FFBDF7BD', fill_type='solid')
    
    
    border_negro = Border(left=Side(style='thin', color='000000'), # negro
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'), 
        bottom=Side(style='thin', color='000000')) 
    
        ### CAMBIO DE CABECERAS     
    ####################################
    
    # Merge cells 
    # numerador y denominador
    ws.merge_cells('B5:C5') 
    ws.merge_cells('D5:E5')

    # intervalo
    ws.merge_cells('D6:E6')

        
    # Combina cela
    ws['B5'] = 'DENOMINADOR'
    ws['D5'] = 'NUMERADOR'
    
    ws['C6'] = 'Atencion Prenatal en el mes de evaluacion'
    ws['D6'] = ' Detección de violencia contra la mujer en gestantes el mismo día de la APN y en el mismo EESS'

    ws['C7'] = 'DX=Z3491 ó Z3492 ó Z3493 ó Z3591 ó Z3592 ó Z3593'
    ws['D7'] = 'DX= 96150.01 + TD = D'
    ws['E7'] = 'DX=96150 + TD= D + LAB = VIF'
    
    ### numerador y denominador 
    ws['B5'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['B5'].font = Font(name = 'Arial', size= 10, bold = True)
    ws['B5'].fill = gray_fill
    ws['B5'].border = border_negro
    
    ws['D5'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['D5'].font = Font(name = 'Arial', size= 10, bold = True)
    ws['D5'].fill = naranja_claro_fill
    ws['D5'].border = border_negro
    
    ### intervalo 
    ws['C6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['C6'].font = Font(name = 'Arial', size= 7)
    ws['C6'].fill = plomo_claro_fill
    ws['C6'].border = border_negro
    
    ws['D6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['D6'].font = Font(name = 'Arial', size= 7)
    ws['D6'].fill = plomo_claro_fill
    ws['D6'].border = border_negro
    
    ws['C7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['C7'].font = Font(name = 'Arial', size= 7)
    ws['C7'].fill = azul_claro_fill
    ws['C7'].border = border_negro
    
    ws['D7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['D7'].font = Font(name = 'Arial', size= 7)
    ws['D7'].fill = azul_claro_fill
    ws['D7'].border = border_negro
    
    ws['E7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['E7'].font = Font(name = 'Arial', size= 7)
    ws['E7'].fill = azul_claro_fill
    ws['E7'].border = border_negro
        
    ws['B6'].alignment = Alignment(horizontal= "center", vertical="center")
    ws['B6'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['B6'].fill = plomo_claro_fill
    ws['B6'].border = border_negro
    ws['B6'] = 'INTERVALO'
    
    ws['B7'].alignment = Alignment(horizontal= "center", vertical="center")
    ws['B7'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['B7'].fill = azul_claro_fill
    ws['B7'].border = border_negro
    ws['B7'] = 'COD HIS'
    
    
    ### BORDE DE CELDAS CONBINADAS
    
    # NUM y DEN
    inicio_columna = 'B'
    fin_columna = 'E'
    fila = 5
    from openpyxl.utils import column_index_from_string
    # Convertir letras de columna a índices numéricos
    indice_inicio = column_index_from_string(inicio_columna)
    indice_fin = column_index_from_string(fin_columna)
    # Iterar sobre las columnas en el rango especificado
    for col in range(indice_inicio, indice_fin + 1):
        celda = ws.cell(row=fila, column=col)
        celda.border = border_negro
    
    # INTERVALO
    inicio_columna = 'B'
    fin_columna = 'E'
    fila = 6
    from openpyxl.utils import column_index_from_string
    # Convertir letras de columna a índices numéricos
    indice_inicio = column_index_from_string(inicio_columna)
    indice_fin = column_index_from_string(fin_columna)
    # Iterar sobre las columnas en el rango especificado
    for col in range(indice_inicio, indice_fin + 1):
        celda = ws.cell(row=fila, column=col)
        celda.border = border_negro
        
    # CODIGO HIS 
    inicio_columna = 'B'
    fin_columna = 'E'
    fila = 7
    from openpyxl.utils import column_index_from_string
    # Convertir letras de columna a índices numéricos
    indice_inicio = column_index_from_string(inicio_columna)
    indice_fin = column_index_from_string(fin_columna)
    # Iterar sobre las columnas en el rango especificado
    for col in range(indice_inicio, indice_fin + 1):
        celda = ws.cell(row=fila, column=col)
        celda.border = border_negro
    
    ##### imprimer fecha y hora del reporte
    fecha_hora_actual = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    nombre_usuario = getpass.getuser()

    # Obtener el usuario actualmente autenticado
    try:
        user = User.objects.get(is_active=True)
    except User.DoesNotExist:
        user = None
    except User.MultipleObjectsReturned:
        # Manejar el caso donde hay múltiples usuarios activos
        user = User.objects.filter(is_active=True).first()  # Por ejemplo, obtener el primero
    # Asignar fecha y hora a la celda A1
    ws['O1'].value = 'Fecha y Hora:'
    ws['P1'].value = fecha_hora_actual

    # Asignar nombre de usuario a la celda A2
    ws['O2'].value = 'Usuario:'
    ws['P2'].value = nombre_usuario
    
    # Formatear las etiquetas en negrita
    etiqueta_font = Font(name='Arial', size=8)
    ws['O1'].font = etiqueta_font
    ws['P1'].font = etiqueta_font
    ws['O2'].font = etiqueta_font
    ws['P2'].font = etiqueta_font

    # Alinear el texto
    ws['O1'].alignment = Alignment(horizontal="right", vertical="center")
    ws['P1'].alignment = Alignment(horizontal="left", vertical="center")
    ws['O2'].alignment = Alignment(horizontal="right", vertical="center")
    ws['P2'].alignment = Alignment(horizontal="left", vertical="center")
    

    ## crea titulo del reporte
    ws['B1'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B1'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['B1'] = 'OFICINA DE TECNOLOGIAS DE LA INFORMACION'
    
    ws['B2'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B2'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['B2'] = 'DIRECCION REGIONAL DE SALUD JUNIN'
    
    ws['B4'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B4'].font = Font(name = 'Arial', size= 12, bold = True)
    ws['B4'] = 'SEGUIMIENTO NOMINAL DEL INDICADOR FICHA VI-01.01: CONDICIÓN PREVIA: PORCENTAJE DE GESTANTES ATENDIDAS EN ESTABLECIMIENTOS DE SALUD, QUE DURANTE EL EMBARAZO LE APLICARON LA FICHA DE DETECCIÓN DE VIOLENCIA CONTRA LA MUJER'
    
    ws['B3'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B3'].font = Font(name = 'Arial', size= 7, color='0000CC')
    ws['B3'] ='El usuario se compromete a mantener la confidencialidad de los datos personales que conozca como resultado del reporte realizado, cumpliendo con lo establecido en la Ley N° 29733 - Ley de Protección de Datos Personales y sus normas complementarias.'
        
    ws['B8'].alignment = Alignment(horizontal= "center", vertical="center")
    ws['B8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['B8'].fill = fill
    ws['B8'].border = border
    ws['B8'] = 'NUM DOC'
        
    ws['C8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['C8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['C8'].fill = fill
    ws['C8'].border = border
    ws['C8'] = 'APN'      
    
    ws['D8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['D8'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
    ws['D8'].fill = blue_fill
    ws['D8'].border = border
    ws['D8'] = 'DETECCION' 
    
    ws['E8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['E8'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
    ws['E8'].fill = blue_fill
    ws['E8'].border = border
    ws['E8'] = 'DETEC + VIF'     
        
    ws['F8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['F8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['F8'].fill = fill
    ws['F8'].border = border
    ws['F8'] = 'MES' 
    
    ws['G8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['G8'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['G8'].fill = gray_fill
    ws['G8'].border = border
    ws['G8'] = 'IND' 
    
    ws['H8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['H8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['H8'].fill = orange_fill
    ws['H8'].border = border
    ws['H8'] = 'UBIGEO'  
    
    ws['I8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['I8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['I8'].fill = orange_fill
    ws['I8'].border = border
    ws['I8'] = 'PROVINCIA'       
    
    ws['J8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['J8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['J8'].fill = orange_fill
    ws['J8'].border = border
    ws['J8'] = 'DISTRITO' 
    
    ws['K8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['K8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['K8'].fill = orange_fill
    ws['K8'].border = border
    ws['K8'] = 'RED'  
    
    ws['L8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['L8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['L8'].fill = orange_fill
    ws['L8'].border = border
    ws['L8'] = 'MICRORED'  
    
    ws['M8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['M8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['M8'].fill = orange_fill
    ws['M8'].border = border
    ws['M8'] = 'COD EST'  
    
    ws['N8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['N8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['N8'].fill = orange_fill
    ws['N8'].border = border
    ws['N8'] = 'ESTABLECIMIENTO'  
    
        
    # Definir estilos
    header_font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    centered_alignment = Alignment(horizontal='center')
    border = Border(left=Side(style='thin', color='A9A9A9'),
            right=Side(style='thin', color='A9A9A9'),
            top=Side(style='thin', color='A9A9A9'),
            bottom=Side(style='thin', color='A9A9A9'))
    header_fill = PatternFill(patternType='solid', fgColor='00B0F0')
    
    
    # Definir los caracteres especiales de check y X
    check_mark = '✓'  # Unicode para check
    x_mark = '✗'  # Unicode para X
    plomo_claro_font = Font(name='Arial', size=7, color='FFEDEDED', bold=False)
    
    # Escribir datos
    for row, record in enumerate(results, start=9):
        for col, value in enumerate(record, start=2):
            cell = ws.cell(row=row, column=col, value=value)

            # Alinear a la izquierda solo en las columnas 6,14,15,16
            if col in [12, 14]:
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='center')
            
            # Aplicar color en la columna 27
            if col == 7:
                if isinstance(value, str):
                    value_upper = value.strip().upper()
                    if value_upper == "NO CUMPLE":
                        cell.fill = PatternFill(patternType='solid', fgColor='FF0000')  # Fondo rojo
                        cell.font = Font(name='Arial', size=7,  bold = True, color='FFFFFF')  # Letra blanca
                    elif value_upper == "CUMPLE":
                        cell.fill = PatternFill(patternType='solid', fgColor='00FF00')  # Fondo verde
                        cell.font = Font(name='Arial', size=7,  bold = True, color='FFFFFF')  # Letra blanca
                    else:
                        cell.font = Font(name='Arial', size=7)
                else:
                    cell.font = Font(name='Arial', size=8)
            
            # Aplicar color de letra en las columnas 7 y 17
            elif col in [22, 45, 46, 61, 63, 74]:
                if isinstance(value, str):
                    value_upper = value.strip().upper()
                    if value_upper == "NO CUMPLE":
                        cell.font = Font(name='Arial', size=7, color="FF0000")  # Letra roja
                    elif value_upper == "CUMPLE":
                        cell.font = Font(name='Arial', size=7, color="00B050")  # Letra verde
                    else:
                        cell.font = Font(name='Arial', size=7)
                else:
                    cell.font = Font(name='Arial', size=7)
            # Fuente normal para otras columnas
            else:
                cell.font = Font(name='Arial', size=8)  # Fuente normal para otras columnas

            # Aplicar caracteres especiales check y X
            if col in [4]:
                if value == 1:
                    cell.value = check_mark  # Insertar check
                    cell.font = Font(name='Arial', size=10, color='00B050')  # Letra verde
                elif value == 0:
                    cell.value = x_mark  # Insertar X
                    cell.font = Font(name='Arial', size=10, color='FF0000')  # Letra roja
                else:
                    cell.font = Font(name='Arial', size=8)  # Fuente normal si no es 1 o 0
                    
                        # Aplicar caracteres especiales check y X
            if col in [5]:
                if value == 1:
                    cell.value = check_mark  # Insertar check
                    cell.font = Font(name='Arial', size=10, color='00B050')  # Letra verde
                elif value == 0:
                    cell.fill = plomo_claro_fill
                    cell.font = plomo_claro_font
                else:
                    cell.font = Font(name='Arial', size=8)  # Fuente normal si no es 1 o 0
            
                        
            cell.border = border

###########################################################################################
# -- COBERTURA PAQUETE NEONATAL
def obtener_cobertura_v1_condicion_previa():
    with connection.cursor() as cursor:
        cursor.execute(
            'SELECT * FROM public."Cobertura_VI0101_CondicionPrevia" ORDER BY "Red", "MicroRed", "Nombre_Establecimiento";'
        )
        return cursor.fetchall()

class RptCoberturaV1CondicionPrevia(TemplateView):
    def get(self, request, *args, **kwargs):
        # Variables ingresadas
                
        # Creación de la consulta
        resultado_cobertura = obtener_cobertura_v1_condicion_previa()
        
        wb = Workbook()
        
        consultas = [
                ('Cobertura', resultado_cobertura)
        ]
        
        for index, (sheet_name, results) in enumerate(consultas):
            if index == 0:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
        
            fill_worksheet_cobertura_v1_condicion_previa(ws, results)
        
        ##########################################################################          
        # Establecer el nombre del archivo
        nombre_archivo = "rpt_cobertura_v1_condicion_previa.xlsx"
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)

        return response


def fill_worksheet_cobertura_v1_condicion_previa(ws, results): 
    # cambia el alto de la columna
    ws.row_dimensions[1].height = 14
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 3
    ws.row_dimensions[4].height = 25
    ws.row_dimensions[5].height = 3
    ws.row_dimensions[7].height = 20
    ws.row_dimensions[8].height = 130
    # cambia el ancho de la columna
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 8
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 8
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 8
    ws.column_dimensions['N'].width = 15
    ws.column_dimensions['O'].width = 15
    ws.column_dimensions['P'].width = 8
    ws.column_dimensions['Q'].width = 15
    ws.column_dimensions['R'].width = 15
    ws.column_dimensions['S'].width = 8
    ws.column_dimensions['T'].width = 15
    ws.column_dimensions['U'].width = 15
    ws.column_dimensions['V'].width = 8
    ws.column_dimensions['W'].width = 15
    ws.column_dimensions['X'].width = 15
    ws.column_dimensions['Y'].width = 8
    ws.column_dimensions['Z'].width = 15    
    ws.column_dimensions['AA'].width = 15
    ws.column_dimensions['AB'].width = 8       
    ws.column_dimensions['AC'].width = 15
    ws.column_dimensions['AD'].width = 15    
    ws.column_dimensions['AE'].width = 8
    ws.column_dimensions['AF'].width = 15    
    ws.column_dimensions['AG'].width = 15
    ws.column_dimensions['AH'].width = 8    
    ws.column_dimensions['AI'].width = 15
    ws.column_dimensions['AJ'].width = 15    
    ws.column_dimensions['AK'].width = 8
    ws.column_dimensions['AL'].width = 15    
    ws.column_dimensions['AM'].width = 15
    ws.column_dimensions['AN'].width = 8    
    ws.column_dimensions['AO'].width = 15
    ws.column_dimensions['AP'].width = 15    
    ws.column_dimensions['AQ'].width = 8    
    
    # linea de division
    ws.freeze_panes = 'H9'
    # Configuración del fondo y el borde
    # Definir el color usando formato aRGB (opacidad completa 'FF' + color RGB)
    fill = PatternFill(start_color='FF60D7E0', end_color='FF60D7E0', fill_type='solid')
    # Definir el color anaranjado usando formato aRGB
    orange_fill = PatternFill(start_color='FFE0A960', end_color='FFE0A960', fill_type='solid')
    # Definir los estilos para gris
    gray_fill = PatternFill(start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid')
    # Definir el estilo de color verde
    green_fill = PatternFill(start_color='FF60E0B3', end_color='FF60E0B3', fill_type='solid')
    # Definir el estilo de color amarillo
    yellow_fill = PatternFill(start_color='FFE0DE60', end_color='FFE0DE60', fill_type='solid')
    # Definir el estilo de color azul
    blue_fill = PatternFill(start_color='FF60A2E0', end_color='FF60A2E0', fill_type='solid')
    # Definir el estilo de color verde 2
    green_fill_2 = PatternFill(start_color='FF60E07E', end_color='FF60E07E', fill_type='solid')
    
    green_font = Font(name='Arial', size=8, color='00FF00')  # Verde
    red_font = Font(name='Arial', size=8, color='FF0000')    # Rojo
    
    
    border = Border(left=Side(style='thin', color='00B0F0'),
                    right=Side(style='thin', color='00B0F0'),
                    top=Side(style='thin', color='00B0F0'),
                    bottom=Side(style='thin', color='00B0F0'))
    
    borde_plomo = Border(left=Side(style='thin', color='A9A9A9'), # Plomo
                    right=Side(style='thin', color='A9A9A9'), # Plomo
                    top=Side(style='thin', color='A9A9A9'), # Plomo
                    bottom=Side(style='thin', color='A9A9A9')) # Plomo
    
    border_negro = Border(left=Side(style='thin', color='000000'), # negro
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'), 
                    bottom=Side(style='thin', color='000000')) 
    
    # Merge cells 
    ws.merge_cells('E7:G7') 
    ws.merge_cells('H7:J7')
    ws.merge_cells('K7:M7')
    ws.merge_cells('N7:P7')
    ws.merge_cells('Q7:S7')
    ws.merge_cells('T7:V7')
    ws.merge_cells('W7:Y7')
    ws.merge_cells('Z7:AB7')
    ws.merge_cells('AC7:AE7')
    ws.merge_cells('AF7:AH7')
    ws.merge_cells('AI7:AK7')
    ws.merge_cells('AL7:AN7')
    ws.merge_cells('AO7:AQ7')

    # Set the value for the merged cell
    ws['E7'] = 'TOTAL'
    ws['H7'] = 'ENERO'
    ws['K7'] = 'FEBRERO'
    ws['N7'] = 'MARZO'
    ws['Q7'] = 'ABRIL'
    ws['T7'] = 'MAYO'
    ws['W7'] = 'JUNIO'
    ws['Z7'] = 'JULIO'
    ws['AC7'] = 'AGOSTO'
    ws['AF7'] = 'SETIEMBRE'
    ws['AI7'] = 'OCTUBRE'
    ws['AL7'] = 'NOVIEMBRE'
    ws['AO7'] = 'DICIEMBRE'

    # Definir el rango desde B3 hasta AA3
    inicio_columna = 'E'
    fin_columna = 'AQ'
    fila = 7
    from openpyxl.utils import column_index_from_string
    # Convertir letras de columna a índices numéricos
    indice_inicio = column_index_from_string(inicio_columna)
    indice_fin = column_index_from_string(fin_columna)

    # Iterar sobre las columnas en el rango especificado
    for col in range(indice_inicio, indice_fin + 1):
        celda = ws.cell(row=fila, column=col)
        celda.border = border_negro
    
    # Definir el rango desde B3 hasta AA3
    inicio_columna_cab = 'B'
    fin_columna_cab = 'AQ'
    fila = 8
    
    # Convertir letras de columna a índices numéricos
    indice_inicio_cab = column_index_from_string(inicio_columna_cab)
    indice_fin_cab = column_index_from_string(fin_columna_cab)

    # Iterar sobre las columnas en el rango especificado
    for col in range(indice_inicio_cab, indice_fin_cab + 1):
        celda = ws.cell(row=fila, column=col)
        celda.border = border_negro
    
    # Apply formatting to the merged cell
    ws['E7'].alignment = Alignment(horizontal="center", vertical="center")
    ws['E7'].font = Font(name='Arial', size=8, bold=True, color='000000')
    ws['E7'].fill = yellow_fill  # Assuming yellow_fill is predefined
    ws['E7'].border = border_negro     # Assuming border is predefined

    ws['H7'].alignment = Alignment(horizontal="center", vertical="center")
    ws['H7'].font = Font(name='Arial', size=8, bold=True, color='000000')
    ws['H7'].fill = blue_fill  # Assuming yellow_fill is predefined
    ws['H7'].border = border_negro 
    
    ws['K7'].alignment = Alignment(horizontal="center", vertical="center")
    ws['K7'].font = Font(name='Arial', size=8, bold=True, color='000000')
    ws['K7'].fill = blue_fill  # Assuming yellow_fill is predefined
    ws['K7'].border = border_negro 
    
    ws['N7'].alignment = Alignment(horizontal="center", vertical="center")
    ws['N7'].font = Font(name='Arial', size=8, bold=True, color='000000')
    ws['N7'].fill = blue_fill  # Assuming yellow_fill is predefined
    ws['N7'].border = border_negro 
    
    ws['Q7'].alignment = Alignment(horizontal="center", vertical="center")
    ws['Q7'].font = Font(name='Arial', size=8, bold=True, color='000000')
    ws['Q7'].fill = blue_fill  # Assuming yellow_fill is predefined
    ws['Q7'].border = border_negro 
    
    ws['T7'].alignment = Alignment(horizontal="center", vertical="center")
    ws['T7'].font = Font(name='Arial', size=8, bold=True, color='000000')
    ws['T7'].fill = blue_fill  # Assuming yellow_fill is predefined
    ws['T7'].border = border_negro 
    
    ws['W7'].alignment = Alignment(horizontal="center", vertical="center")
    ws['W7'].font = Font(name='Arial', size=8, bold=True, color='000000')
    ws['W7'].fill = blue_fill  # Assuming yellow_fill is predefined
    ws['W7'].border = border_negro 
    
    ws['Z7'].alignment = Alignment(horizontal="center", vertical="center")
    ws['Z7'].font = Font(name='Arial', size=8, bold=True, color='000000')
    ws['Z7'].fill = blue_fill  # Assuming yellow_fill is predefined
    ws['Z7'].border = border_negro 
    
    ws['AC7'].alignment = Alignment(horizontal="center", vertical="center")
    ws['AC7'].font = Font(name='Arial', size=8, bold=True, color='000000')
    ws['AC7'].fill = blue_fill  # Assuming yellow_fill is predefined
    ws['AC7'].border = border_negro 
    
    ws['AF7'].alignment = Alignment(horizontal="center", vertical="center")
    ws['AF7'].font = Font(name='Arial', size=8, bold=True, color='000000')
    ws['AF7'].fill = blue_fill  # Assuming yellow_fill is predefined
    ws['AF7'].border = border_negro 
    
    ws['AI7'].alignment = Alignment(horizontal="center", vertical="center")
    ws['AI7'].font = Font(name='Arial', size=8, bold=True, color='000000')
    ws['AI7'].fill = blue_fill  # Assuming yellow_fill is predefined
    ws['AI7'].border = border_negro 
    
    ws['AL7'].alignment = Alignment(horizontal="center", vertical="center")
    ws['AL7'].font = Font(name='Arial', size=8, bold=True, color='000000')
    ws['AL7'].fill = blue_fill  # Assuming yellow_fill is predefined
    ws['AL7'].border = border_negro 
    
    ws['AO7'].alignment = Alignment(horizontal="center", vertical="center")
    ws['AO7'].font = Font(name='Arial', size=8, bold=True, color='000000')
    ws['AO7'].fill = blue_fill  # Assuming yellow_fill is predefined
    ws['AO7'].border = border_negro 
    
    ## crea titulo del reporte
    ws['B1'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B1'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['B1'] = 'OFICINA DE TECNOLOGIAS DE LA INFORMACION'
    
    ws['B2'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B2'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['B2'] = 'DIRECCION REGIONAL DE SALUD JUNIN'
    
    ws['B4'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B4'].font = Font(name = 'Arial', size= 12, bold = True)
    ws['B4'] = 'COBERURA DEL INDICADOR VI-01.01. CONDICIÓN PREVIA: 95% PORCENTAJE DE GESTANTES ATENDIDAS EN ESTABLECIMIENTOS DE SALUD, QUE DURANTE EL EMBARAZO LE APLICARON LA FICHA DE DETECCIÓN DE VIOLENCIA CONTRA LA MUJER'
    
    ws['B6'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B6'].font = Font(name = 'Arial', size= 7, bold = True, color='0000CC')
    ws['B6'] ='NOTAS: Primera columna "NUMERADOR", segunda columna "DENOMINADOR", tercera columna "PORCENTAJE AVANCE"'
        
    ws['B8'].alignment = Alignment(horizontal= "center", vertical="center")
    ws['B8'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['B8'].fill = yellow_fill
    ws['B8'].border = border_negro
    ws['B8'] = 'RED'
    
    ws['C8'].alignment = Alignment(horizontal= "center", vertical="center")
    ws['C8'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['C8'].fill = yellow_fill
    ws['C8'].border = border_negro
    ws['C8'] = 'MICRORED'
    
    ws['D8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['D8'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['D8'].fill = yellow_fill
    ws['D8'].border = border_negro
    ws['D8'] = 'ESTABLECIMIENTO'      
    
    ws['E8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['E8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['E8'].fill = yellow_fill
    ws['E8'].border = border_negro
    ws['E8'] = 'N° de gestantes del denominador a quienes le aplicaron la ficha de detección de violencia contra la mujer (tamizaje), registradas en el HIS'
    
    ws['F8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['F8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['F8'].fill = yellow_fill
    ws['F8'].border = border_negro
    ws['F8'] = 'N° de gestantes atendidas durante el embarazo, en Establecimientos de Salud del primer nivel y segundo nivel de atención, registrados en el HIS'
    
    ws['G8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['G8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['G8'].fill = yellow_fill
    ws['G8'].border = border_negro
    ws['G8'] = '% Avance (Num/Den)'    
    
    ws['H8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['H8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['H8'].fill = blue_fill
    ws['H8'].border = border_negro
    ws['H8'] = 'N° de gestantes del denominador a quienes le aplicaron la ficha de detección de violencia contra la mujer (tamizaje), registradas en el HIS'    
    
    ws['I8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['I8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['I8'].fill = blue_fill
    ws['I8'].border = border_negro
    ws['I8'] = 'N° de gestantes atendidas durante el embarazo, en Establecimientos de Salud del primer nivel y segundo nivel de atención, registrados en el HIS' 
    
    ws['J8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['J8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['J8'].fill = gray_fill
    ws['J8'].border = border_negro
    ws['J8'] = '% Avance (Num/Den)'    
    
    ws['K8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['K8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['K8'].fill = blue_fill
    ws['K8'].border = border_negro
    ws['K8'] = 'N° de gestantes del denominador a quienes le aplicaron la ficha de detección de violencia contra la mujer (tamizaje), registradas en el HIS'     
    
    ws['L8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['L8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['L8'].fill = blue_fill
    ws['L8'].border = border_negro
    ws['L8'] = 'N° de gestantes atendidas durante el embarazo, en Establecimientos de Salud del primer nivel y segundo nivel de atención, registrados en el HIS' 
    
    ws['M8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['M8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['M8'].fill = gray_fill
    ws['M8'].border = border_negro
    ws['M8'] = '% Avance (Num/Den)'
    
    ws['N8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['N8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['N8'].fill = blue_fill
    ws['N8'].border = border_negro
    ws['N8'] = 'N° de gestantes del denominador a quienes le aplicaron la ficha de detección de violencia contra la mujer (tamizaje), registradas en el HIS'   
    
    ws['O8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['O8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['O8'].fill = blue_fill
    ws['O8'].border = border_negro
    ws['O8'] = 'N° de gestantes atendidas durante el embarazo, en Establecimientos de Salud del primer nivel y segundo nivel de atención, registrados en el HIS'
    
    ws['P8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['P8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['P8'].fill = gray_fill
    ws['P8'].border = border_negro
    ws['P8'] = '% Avance (Num/Den)'     
    
    ws['Q8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['Q8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['Q8'].fill = blue_fill
    ws['Q8'].border = border_negro
    ws['Q8'] = 'N° de gestantes del denominador a quienes le aplicaron la ficha de detección de violencia contra la mujer (tamizaje), registradas en el HIS'   
    
    ws['R8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['R8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['R8'].fill = blue_fill
    ws['R8'].border = border_negro
    ws['R8'] = 'N° de gestantes atendidas durante el embarazo, en Establecimientos de Salud del primer nivel y segundo nivel de atención, registrados en el HIS' 
    
    ws['S8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['S8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['S8'].fill = gray_fill
    ws['S8'].border = border_negro
    ws['S8'] = '% Avance (Num/Den)'    
    
    ws['T8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['T8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['T8'].fill = blue_fill
    ws['T8'].border = border_negro
    ws['T8'] = 'N° de gestantes del denominador a quienes le aplicaron la ficha de detección de violencia contra la mujer (tamizaje), registradas en el HIS'    
    
    ws['U8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['U8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['U8'].fill = blue_fill
    ws['U8'].border = border_negro
    ws['U8'] = 'N° de gestantes atendidas durante el embarazo, en Establecimientos de Salud del primer nivel y segundo nivel de atención, registrados en el HIS'
    
    ws['V8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['V8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['V8'].fill = gray_fill
    ws['V8'].border = border_negro
    ws['V8'] = '% Avance (Num/Den)'    
    
    ws['W8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['W8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['W8'].fill = blue_fill
    ws['W8'].border = border_negro
    ws['W8'] = 'N° de gestantes del denominador a quienes le aplicaron la ficha de detección de violencia contra la mujer (tamizaje), registradas en el HIS'   
        
    ws['X8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['X8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['X8'].fill = blue_fill
    ws['X8'].border = border_negro
    ws['X8'] = 'N° de gestantes atendidas durante el embarazo, en Establecimientos de Salud del primer nivel y segundo nivel de atención, registrados en el HIS'

    ws['Y8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['Y8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['Y8'].fill = gray_fill
    ws['Y8'].border = border_negro
    ws['Y8'] = '% Avance (Num/Den)'    
    
    ws['Z8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['Z8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['Z8'].fill = blue_fill
    ws['Z8'].border = border_negro
    ws['Z8'] = 'N° de gestantes del denominador a quienes le aplicaron la ficha de detección de violencia contra la mujer (tamizaje), registradas en el HIS'   

    ws['AA8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AA8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['AA8'].fill = blue_fill
    ws['AA8'].border = border_negro
    ws['AA8'] = 'N° de gestantes atendidas durante el embarazo, en Establecimientos de Salud del primer nivel y segundo nivel de atención, registrados en el HIS'
    
    ws['AB8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AB8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['AB8'].fill = gray_fill
    ws['AB8'].border = border_negro
    ws['AB8'] = '% Avance (Num/Den)'    
    
    ws['AC8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AC8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['AC8'].fill = blue_fill
    ws['AC8'].border = border_negro
    ws['AC8'] = 'N° de gestantes del denominador a quienes le aplicaron la ficha de detección de violencia contra la mujer (tamizaje), registradas en el HIS'   
    
    ws['AD8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AD8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['AD8'].fill = blue_fill
    ws['AD8'].border = border_negro
    ws['AD8'] = 'N° de gestantes atendidas durante el embarazo, en Establecimientos de Salud del primer nivel y segundo nivel de atención, registrados en el HIS'
    
    ws['AE8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AE8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['AE8'].fill = gray_fill
    ws['AE8'].border = border_negro
    ws['AE8'] = '% Avance (Num/Den)'    
    
    ws['AF8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AF8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['AF8'].fill = blue_fill
    ws['AF8'].border = border_negro
    ws['AF8'] = 'N° de gestantes del denominador a quienes le aplicaron la ficha de detección de violencia contra la mujer (tamizaje), registradas en el HIS'   
    
    ws['AG8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AG8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['AG8'].fill = blue_fill
    ws['AG8'].border = border_negro
    ws['AG8'] = 'N° de gestantes atendidas durante el embarazo, en Establecimientos de Salud del primer nivel y segundo nivel de atención, registrados en el HIS' 
    
    ws['AH8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AH8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['AH8'].fill = gray_fill
    ws['AH8'].border = border_negro
    ws['AH8'] = '% Avance (Num/Den)'    
    
    ws['AI8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AI8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['AI8'].fill = blue_fill
    ws['AI8'].border = border_negro
    ws['AI8'] = 'N° de gestantes del denominador a quienes le aplicaron la ficha de detección de violencia contra la mujer (tamizaje), registradas en el HIS'    
    
    ws['AJ8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AJ8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['AJ8'].fill = blue_fill
    ws['AJ8'].border = border_negro
    ws['AJ8'] = 'N° de gestantes atendidas durante el embarazo, en Establecimientos de Salud del primer nivel y segundo nivel de atención, registrados en el HIS'
    
    ws['AK8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AK8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['AK8'].fill = gray_fill
    ws['AK8'].border = border_negro
    ws['AK8'] = '% Avance (Num/Den)'    
    
    ws['AL8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AL8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['AL8'].fill = blue_fill
    ws['AL8'].border = border_negro
    ws['AL8'] = 'N° de gestantes del denominador a quienes le aplicaron la ficha de detección de violencia contra la mujer (tamizaje), registradas en el HIS'    
    
    ws['AM8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AM8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['AM8'].fill = blue_fill
    ws['AM8'].border = border_negro
    ws['AM8'] = 'N° de gestantes atendidas durante el embarazo, en Establecimientos de Salud del primer nivel y segundo nivel de atención, registrados en el HIS'
    
    ws['AN8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AN8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['AN8'].fill = gray_fill
    ws['AN8'].border = border_negro
    ws['AN8'] = '% Avance (Num/Den)'    
    
    ws['AO8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AO8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['AO8'].fill = blue_fill
    ws['AO8'].border = border_negro
    ws['AO8'] = 'N° de gestantes del denominador a quienes le aplicaron la ficha de detección de violencia contra la mujer (tamizaje), registradas en el HIS'   
    
    ws['AP8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AP8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['AP8'].fill = blue_fill
    ws['AP8'].border = border_negro
    ws['AP8'] = 'N° de gestantes atendidas durante el embarazo, en Establecimientos de Salud del primer nivel y segundo nivel de atención, registrados en el HIS' 
    
    ws['AQ8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AQ8'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['AQ8'].fill = gray_fill
    ws['AQ8'].border = border_negro
    ws['AQ8'] = '% Avance (Num/Den)'    
    
    # Definir estilos
    header_font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    centered_alignment = Alignment(horizontal='center')
    border = Border(left=Side(style='thin', color='A9A9A9'),
            right=Side(style='thin', color='A9A9A9'),
            top=Side(style='thin', color='A9A9A9'),
            bottom=Side(style='thin', color='A9A9A9'))
    header_fill = PatternFill(patternType='solid', fgColor='00B0F0')
    
    # Definir los caracteres especiales de check y X
    
    # Escribir datos
    for row, record in enumerate(results, start=9):
        for col, value in enumerate(record, start=2):
            cell = ws.cell(row=row, column=col, value=value)

            # Alinear a la izquierda solo en las columnas 6,14,15,16
            if col in [2, 3, 4]:
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='center')

            # Aplicar formato basado en el valor para columnas específicas
            if col in [7, 10, 13, 16, 19, 22, 25, 28, 31, 34, 37, 40, 43]:
                try:
                    value_float = float(value)
                except (ValueError, TypeError):
                    # Si el valor no es numérico, aplicar formato por defecto
                    cell.font = Font(name='Arial', size=7)
                else:
                    # Si los valores están entre 0 y 100, dividimos entre 100
                    if value_float > 1:
                        value_float = value_float / 100
                        cell.value = value_float

                    # Establecer el formato de número a porcentaje
                    cell.number_format = '0.0%'

                    if value_float >= 1.00:
                        # Colorear la celda de verde
                        cell.fill = PatternFill(patternType='solid', fgColor='00B050')  # Fondo verde
                        cell.font = Font(name='Arial', size=7, color='000000')  # Letra negra
                    else:
                        # Colorear la celda de rojo con letras blancas
                        cell.fill = PatternFill(patternType='solid', fgColor='FF0000')  # Fondo rojo
                        cell.font = Font(name='Arial', size=7, color='FFFFFF')  # Letra blanca
            # Fuente normal para otras columnas
            else:
                cell.font = Font(name='Arial', size=8)
            
            cell.border = border