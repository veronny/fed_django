# app_name/styles.py

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# Fuentes
FONT_ARIAL_7_BOLD = Font(name='Arial', size=7, bold=True)
FONT_ARIAL_8 = Font(name='Arial', size=8)
FONT_ARIAL_8_BOLD_WHITE = Font(name='Arial', size=8, bold=True, color='FFFFFF')
FONT_ARIAL_7_BOLD_BLUE = Font(name='Arial', size=7, bold=True, color='0000CC')
FONT_ARIAL_7_WHITE_BOLD = Font(name='Arial', size=7, bold=True, color='FFFFFF')
FONT_ARIAL_12_BOLD = Font(name='Arial', size=12, bold=True)

# Colores de relleno
FILL_CYAN = PatternFill(start_color='FF60D7E0', end_color='FF60D7E0', fill_type='solid')
FILL_ORANGE = PatternFill(start_color='FFE0A960', end_color='FFE0A960', fill_type='solid')
FILL_GRAY = PatternFill(start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid')
FILL_GREEN = PatternFill(start_color='FF60E0B3', end_color='FF60E0B3', fill_type='solid')
FILL_YELLOW = PatternFill(start_color='FFE0DE60', end_color='FFE0DE60', fill_type='solid')
FILL_BLUE = PatternFill(start_color='FF60A2E0', end_color='FF60A2E0', fill_type='solid')
FILL_GREEN_2 = PatternFill(start_color='FF60E07E', end_color='FF60E07E', fill_type='solid')

# Bordes
BORDER_THIN_GRAY = Border(
    left=Side(style='thin', color='A9A9A9'),
    right=Side(style='thin', color='A9A9A9'),
    top=Side(style='thin', color='A9A9A9'),
    bottom=Side(style='thin', color='A9A9A9')
)

# Alineaciones
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_LEFT = Alignment(horizontal='left', vertical='center')

# Caracteres especiales
CHECK_MARK = '✓'
X_MARK = '✗'
SUB_CUMPLE = 'CUMPLE'
SUB_NO_CUMPLE = 'NO CUMPLE'