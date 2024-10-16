from django.shortcuts import render

# TABLERO Adolescente Hemopglobina 
from django.db import connection
from django.http import JsonResponse
from base.models import MAESTRO_HIS_ESTABLECIMIENTO, DimPeriodo
from django.db.models.functions import Substr
import logging

# report excel
from django.http.response import HttpResponse
from django.views.generic.base import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import openpyxl
from openpyxl.utils import get_column_letter

from django.db.models.functions import Substr

from datetime import datetime
import locale

from django.db.models import IntegerField               # Importar IntegerField
from django.db.models.functions import Cast, Substr     # Importar Cast y Substr


logger = logging.getLogger(__name__)
# Create your views here.
def obtener_distritos(provincia):
    distritos = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Provincia=provincia).values('Distrito').distinct().order_by('Distrito')
    return list(distritos)

def obtener_avance_v2_tamizaje_violencia(red):
    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT DISTINCT * FROM public.obtener_avance_v2_tamizaje_violencia(%s)",
            [red]
        )
        return cursor.fetchall()

def obtener_ranking_v2_tamizaje_violencia(anio, mes):
    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT DISTINCT * FROM public.obtener_ranking_v2_tamizaje_violencia(%s, %s)",
            [anio, mes]
        )
        result = cursor.fetchall()
        return result
    
def index_v2_tamizaje_violencia(request):
    # RANKING 
    anio = request.GET.get('anio')  # Valor predeterminado# Valor predeterminado
    mes_seleccionado = request.GET.get('mes')
    # GRAFICO
    red_seleccionada = request.GET.get('red')
    red = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Disa='JUNIN').values_list('Red', flat=True).distinct().order_by('Red')
    # Si la solicitud es AJAX
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            # Obtener datos de RANKING 
            resultados_ranking_obtener_v2_tamizaje_violencia = obtener_ranking_v2_tamizaje_violencia(anio,mes_seleccionado)
            # Obtener datos de AVANCE GRAFICO MESES
            resultados_avance_obtener_v2_tamizaje_violencia = obtener_avance_v2_tamizaje_violencia(red_seleccionada)
            
            # Procesar los resultados
            if any(len(row) < 4 for row in resultados_ranking_obtener_v2_tamizaje_violencia):
                raise ValueError("Algunas filas del ranking no tienen suficientes elementos")
            
            data = {               
                #ranking
                'red': [],
                'num_r': [],
                'den_r': [],
                'avance_r': [],
                
                #avance meses
                'mes': [],
                'num': [],
                'den': [],
                'avance': [],
            }
            #RANKING
            for index, row in enumerate(resultados_ranking_obtener_v2_tamizaje_violencia):
                try:
                    # Verifica que la tupla tenga exactamente 4 elementos
                    if len(row) != 4:
                        raise ValueError(f"La fila {index} no tiene 4 elementos: {row}")

                    red_value = row[0] if row[0] is not None else ''
                    num_r_value = float(row[1]) if row[1] is not None else 0.0
                    den_r_value = float(row[2]) if row[2] is not None else 0.0
                    avance_r_value = float(row[3]) if row[3] is not None else 0.0

                    data['red'].append(red_value)
                    data['num_r'].append(num_r_value)
                    data['den_r'].append(den_r_value)
                    data['avance_r'].append(avance_r_value)

                except Exception as e:
                    logger.error(f"Error procesando la fila {index}: {str(e)}")
            
            #AVANCE GRAFICO MESES
            for index, row in enumerate(resultados_avance_obtener_v2_tamizaje_violencia):
                try:
                    # Verifica que la tupla tenga exactamente 4 elementos
                    if len(row) != 5:
                        raise ValueError(f"La fila {index} no tiene 5 elementos: {row}")

                    mes_value = row[1] if row[1] is not None else ''
                    num_value = float(row[2]) if row[2] is not None else 0.0
                    den_value = float(row[3]) if row[3] is not None else 0.0
                    avance_value = float(row[4]) if row[4] is not None else 0.0

                    data['mes'].append(mes_value)
                    data['num'].append(num_value)
                    data['den'].append(den_value)
                    data['avance'].append(avance_value)

                except Exception as e:
                    logger.error(f"Error procesando la fila {index}: {str(e)}")
            
            return JsonResponse(data)

        except Exception as e:
            logger.error(f"Error al obtener datos: {str(e)}")

    # Si no es una solicitud AJAX, renderiza la página principal
    return render(request, 'v2_tamizaje_violencia/index_v2_tamizaje_violencia.html', {
        'red': red,
        'mes_seleccionado': mes_seleccionado,
    })

## SEGUIMIENTO
def get_redes_v2_tamizaje_violencia(request,redes_id):
    redes = (
            MAESTRO_HIS_ESTABLECIMIENTO
            .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL',Departamento='JUNIN')
            .annotate(codigo_red_filtrado=Substr('Codigo_Red', 1, 4))
            .values('Red','codigo_red_filtrado')
            .distinct()
            .order_by('Red')
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'redes': redes,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
    }
    
    return render(request, 'v2_tamizaje_violencia/redes.html', context)

def obtener_seguimiento_redes_v2_tamizaje_violencia(p_red,p_inicio,p_fin):
    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT DISTINCT * FROM public.fn_seguimiento_v2_tamizaje_violencia(%s, %s, %s)",
            [p_red, p_inicio, p_fin]
        )
        return cursor.fetchall()

class RptV1CondicionPreviaRed(TemplateView):
    def get(self, request, *args, **kwargs):
        # Variables ingresadas
        p_red = request.GET.get('red')
        p_inicio = request.GET.get('fecha_inicio')
        p_fin = request.GET.get('fecha_fin')
        # Creación de la consulta
        resultado_seguimiento = obtener_seguimiento_redes_v2_tamizaje_violencia(p_red, p_inicio, p_fin)
        
        wb = Workbook()
        
        consultas = [
                ('Seguimiento', resultado_seguimiento)
        ]
        
        for index, (sheet_name, results) in enumerate(consultas):
            if index == 0:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
        
            fill_worksheet_v2_tamizaje_violencia(ws, results)
        ##########################################################################          
        # Establecer el nombre del archivo
        nombre_archivo = "rpt_v2_tamizaje_violencia.xlsx"
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)

        return response

def fill_worksheet_v2_tamizaje_violencia(ws, results): 
    # cambia el alto de la columna
    ws.row_dimensions[1].height = 14
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 3
    ws.row_dimensions[4].height = 25
    ws.row_dimensions[5].height = 3
    ws.row_dimensions[7].height = 3
    ws.row_dimensions[8].height = 25
    # cambia el ancho de la columna
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 9
    ws.column_dimensions['C'].width = 9
    ws.column_dimensions['D'].width = 6
    ws.column_dimensions['E'].width = 6
    ws.column_dimensions['F'].width = 6
    ws.column_dimensions['G'].width = 6
    ws.column_dimensions['H'].width = 6
    ws.column_dimensions['I'].width = 9
    ws.column_dimensions['J'].width = 10
    
    
    ws.column_dimensions['K'].width = 9
    ws.column_dimensions['L'].width = 16    
    ws.column_dimensions['M'].width = 20
    ws.column_dimensions['N'].width = 16
    ws.column_dimensions['O'].width = 20
    ws.column_dimensions['P'].width = 9
    ws.column_dimensions['Q'].width = 30

    # linea de division
    ws.freeze_panes = 'E9'
    # Configuración del fondo y el borde
    # Definir el color usando formato aRGB (opacidad completa 'FF' + color RGB)
    fill = PatternFill(start_color='FF60D7E0', end_color='FF60D7E0', fill_type='solid')
    # Definir el color anaranjado usando formato aRGB
    orange_fill = PatternFill(start_color='FFE0A960', end_color='FFE0A960', fill_type='solid')
    # Definir los estilos para gris
    gray_fill = PatternFill(start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid')
    # Definir el estilo de color verde
    green_fill = PatternFill(start_color='FF60E0B3', end_color='FF60E0B3', fill_type='solid')
    # Definir el estilo de color amarillo
    yellow_fill = PatternFill(start_color='FFE0DE60', end_color='FFE0DE60', fill_type='solid')
    # Definir el estilo de color azul
    blue_fill = PatternFill(start_color='FF60A2E0', end_color='FF60A2E0', fill_type='solid')
    # Definir el estilo de color verde 2
    green_fill_2 = PatternFill(start_color='FF60E07E', end_color='FF60E07E', fill_type='solid')   
    
    green_font = Font(name='Arial', size=8, color='00FF00')  # Verde
    red_font = Font(name='Arial', size=8, color='FF0000')    # Rojo
    
    border = Border(left=Side(style='thin', color='00B0F0'),
                    right=Side(style='thin', color='00B0F0'),
                    top=Side(style='thin', color='00B0F0'),
                    bottom=Side(style='thin', color='00B0F0'))
    borde_plomo = Border(left=Side(style='thin', color='A9A9A9'), # Plomo
                    right=Side(style='thin', color='A9A9A9'), # Plomo
                    top=Side(style='thin', color='A9A9A9'), # Plomo
                    bottom=Side(style='thin', color='A9A9A9')) # Plomo
    
    ## crea titulo del reporte
    ws['B1'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B1'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['B1'] = 'OFICINA DE TECNOLOGIAS DE LA INFORMACION'
    
    ws['B2'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B2'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['B2'] = 'DIRECCION REGIONAL DE SALUD JUNIN'
    
    ws['B4'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B4'].font = Font(name = 'Arial', size= 12, bold = True)
    ws['B4'] = 'SEGUIMIENTO NOMINAL DEL INDICADOR FICHA VI-01.02: PORCENTAJE DE GESTANTES ATENDIDAS EN ESTABLECIMIENTOS DE SALUD DEL PRIMER Y SEGUNDO NIVEL DE ATENCIÓN DEL GOBIERNO REGIONAL, QUE CUENTAN TAMIZAJE POSITIVO DE VIOLENCIA CONTRA LA MUJER'
    
    ws['B6'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B6'].font = Font(name = 'Arial', size= 7, bold = True, color='0000CC')
    ws['B6'] ='El usuario se compromete a mantener la confidencialidad de los datos personales que conozca como resultado del reporte realizado, cumpliendo con lo establecido en la Ley N° 29733 - Ley de Protección de Datos Personales y sus normas complementarias.'
        
    ws['B8'].alignment = Alignment(horizontal= "center", vertical="center")
    ws['B8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['B8'].fill = fill
    ws['B8'].border = border
    ws['B8'] = 'NUM DOC'
        
    ws['C8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['C8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['C8'].fill = fill
    ws['C8'].border = border
    ws['C8'] = 'APN'      
    
    ws['D8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['D8'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
    ws['D8'].fill = blue_fill
    ws['D8'].border = border
    ws['D8'] = 'VAL APN' 
    
    ws['E8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['E8'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
    ws['E8'].fill = yellow_fill
    ws['E8'].border = border
    ws['E8'] = 'VAL VIO_1'     
    
    ws['F8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['F8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['F8'].fill = fill
    ws['F8'].border = border
    ws['F8'] = 'VAL VID_2'      
    
    ws['G8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['G8'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
    ws['G8'].fill = blue_fill
    ws['G8'].border = border
    ws['G8'] = 'RX VIO_1' 
    
    ws['H8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['H8'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
    ws['H8'].fill = yellow_fill
    ws['H8'].border = border
    ws['H8'] = 'RX VIO_2'     
        
    ws['I8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['I8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['I8'].fill = fill
    ws['I8'].border = border
    ws['I8'] = 'MES' 
    
    ws['J8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['J8'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['J8'].fill = gray_fill
    ws['J8'].border = border
    ws['J8'] = 'IND' 
    
    ws['K8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['K8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['K8'].fill = orange_fill
    ws['K8'].border = border
    ws['K8'] = 'UBIGEO'  
    
    ws['L8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['L8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['L8'].fill = orange_fill
    ws['L8'].border = border
    ws['L8'] = 'PROVINCIA'       
    
    ws['M8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['M8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['M8'].fill = orange_fill
    ws['M8'].border = border
    ws['M8'] = 'DISTRITO' 
    
    ws['N8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['N8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['N8'].fill = orange_fill
    ws['N8'].border = border
    ws['N8'] = 'RED'  
    
    ws['O8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['O8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['O8'].fill = orange_fill
    ws['O8'].border = border
    ws['O8'] = 'MICRORED'  
    
    ws['P8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['P8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['P8'].fill = orange_fill
    ws['P8'].border = border
    ws['P8'] = 'COD EST'  
    
    ws['Q8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['Q8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['Q8'].fill = orange_fill
    ws['Q8'].border = border
    ws['Q8'] = 'ESTABLECIMIENTO'  
    
        
    # Definir estilos
    header_font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    centered_alignment = Alignment(horizontal='center')
    border = Border(left=Side(style='thin', color='A9A9A9'),
            right=Side(style='thin', color='A9A9A9'),
            top=Side(style='thin', color='A9A9A9'),
            bottom=Side(style='thin', color='A9A9A9'))
    header_fill = PatternFill(patternType='solid', fgColor='00B0F0')
    
    
    # Definir los caracteres especiales de check y X
    check_mark = '✓'  # Unicode para check
    x_mark = '✗'  # Unicode para X
    
    
    # Escribir datos
    for row, record in enumerate(results, start=9):
        for col, value in enumerate(record, start=2):
            cell = ws.cell(row=row, column=col, value=value)

            # Alinear a la izquierda solo en las columnas 6,14,15,16
            if col in [12, 14]:
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='center')

            # Aplicar color en la columna 27
            if col == 10:
                if isinstance(value, str):
                    value_upper = value.strip().upper()
                    if value_upper == "NO CUMPLE":
                        cell.fill = PatternFill(patternType='solid', fgColor='FF0000')  # Fondo rojo
                        cell.font = Font(name='Arial', size=7,  bold = True, color='FFFFFF')  # Letra blanca
                    elif value_upper == "CUMPLE":
                        cell.fill = PatternFill(patternType='solid', fgColor='00FF00')  # Fondo verde
                        cell.font = Font(name='Arial', size=7,  bold = True, color='FFFFFF')  # Letra blanca
                    else:
                        cell.font = Font(name='Arial', size=7)
                else:
                    cell.font = Font(name='Arial', size=8)
            
            # Aplicar color de letra en las columnas 7 y 17
            elif col in [22, 45, 46, 61, 63, 74]:
                if isinstance(value, str):
                    value_upper = value.strip().upper()
                    if value_upper == "NO CUMPLE":
                        cell.font = Font(name='Arial', size=7, color="FF0000")  # Letra roja
                    elif value_upper == "CUMPLE":
                        cell.font = Font(name='Arial', size=7, color="00B050")  # Letra verde
                    else:
                        cell.font = Font(name='Arial', size=7)
                else:
                    cell.font = Font(name='Arial', size=7)
            # Fuente normal para otras columnas
            else:
                cell.font = Font(name='Arial', size=8)  # Fuente normal para otras columnas

            # Aplicar caracteres especiales check y X
            if col in [4,5,6,7,8]:
                if value == 1:
                    cell.value = check_mark  # Insertar check
                    cell.font = Font(name='Arial', size=10, color='00B050')  # Letra verde
                elif value == 0:
                    cell.value = x_mark  # Insertar X
                    cell.font = Font(name='Arial', size=10, color='FF0000')  # Letra roja
                else:
                    cell.font = Font(name='Arial', size=8)  # Fuente normal si no es 1 o 0
            
                        
            cell.border = border