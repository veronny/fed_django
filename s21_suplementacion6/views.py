from django.shortcuts import render

# TABLERO Adolescente Hemopglobina 
from django.db import connection
from django.http import JsonResponse
from base.models import MAESTRO_HIS_ESTABLECIMIENTO, DimPeriodo
from django.db.models.functions import Substr
import logging

# report excel
from django.http.response import HttpResponse
from django.views.generic.base import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import openpyxl
from openpyxl.utils import get_column_letter

from django.db.models.functions import Substr

from datetime import datetime
import locale

from django.db.models import IntegerField  # Importar IntegerField
from django.db.models.functions import Cast, Substr  # Importar Cast y Substr


logger = logging.getLogger(__name__)
# Create your views here.
def obtener_distritos(provincia):
    distritos = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Provincia=provincia).values('Distrito').distinct().order_by('Distrito')
    return list(distritos)

def obtener_avance_s21_suplementacion6(red):
    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT DISTINCT * FROM public.obtener_avance_s21_suplementacion6(%s)",
            [red]
        )
        return cursor.fetchall()

def obtener_ranking_s21_suplementacion6(anio, mes):
    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT DISTINCT * FROM public.obtener_ranking_s21_suplementacion6(%s, %s)",
            [anio, mes]
        )
        result = cursor.fetchall()
        return result
    
def index_s21_suplementacion6(request):
    # RANKING 
    anio = request.GET.get('anio')  # Valor predeterminado# Valor predeterminado
    mes_seleccionado = request.GET.get('mes')
    # GRAFICO
    red_seleccionada = request.GET.get('red')
    red = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Disa='JUNIN').values_list('Red', flat=True).distinct().order_by('Red')
    # Si la solicitud es AJAX
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            # Obtener datos de RANKING 
            resultados_ranking_obtener_s21_suplementacion6 = obtener_ranking_s21_suplementacion6(anio,mes_seleccionado)
            # Obtener datos de AVANCE GRAFICO MESES
            resultados_avance_obtener_s21_suplementacion6 = obtener_avance_s21_suplementacion6(red_seleccionada)
            
            # Procesar los resultados
            if any(len(row) < 4 for row in resultados_ranking_obtener_s21_suplementacion6):
                raise ValueError("Algunas filas del ranking no tienen suficientes elementos")
            
            data = {               
                #ranking
                'red': [],
                'num_r': [],
                'den_r': [],
                'avance_r': [],
                
                #avance meses
                'mes': [],
                'num': [],
                'den': [],
                'avance': [],
            }
            #RANKING
            for index, row in enumerate(resultados_ranking_obtener_s21_suplementacion6):
                try:
                    # Verifica que la tupla tenga exactamente 4 elementos
                    if len(row) != 4:
                        raise ValueError(f"La fila {index} no tiene 4 elementos: {row}")

                    red_value = row[0] if row[0] is not None else ''
                    num_r_value = float(row[1]) if row[1] is not None else 0.0
                    den_r_value = float(row[2]) if row[2] is not None else 0.0
                    avance_r_value = float(row[3]) if row[3] is not None else 0.0

                    data['red'].append(red_value)
                    data['num_r'].append(num_r_value)
                    data['den_r'].append(den_r_value)
                    data['avance_r'].append(avance_r_value)

                except Exception as e:
                    logger.error(f"Error procesando la fila {index}: {str(e)}")
            
            #AVANCE GRAFICO MESES
            for index, row in enumerate(resultados_avance_obtener_s21_suplementacion6):
                try:
                    # Verifica que la tupla tenga exactamente 4 elementos
                    if len(row) != 5:
                        raise ValueError(f"La fila {index} no tiene 5 elementos: {row}")

                    mes_value = row[1] if row[1] is not None else ''
                    num_value = float(row[2]) if row[2] is not None else 0.0
                    den_value = float(row[3]) if row[3] is not None else 0.0
                    avance_value = float(row[4]) if row[4] is not None else 0.0

                    data['mes'].append(mes_value)
                    data['num'].append(num_value)
                    data['den'].append(den_value)
                    data['avance'].append(avance_value)

                except Exception as e:
                    logger.error(f"Error procesando la fila {index}: {str(e)}")
            
            return JsonResponse(data)

        except Exception as e:
            logger.error(f"Error al obtener datos: {str(e)}")

    # Si no es una solicitud AJAX, renderiza la página principal
    return render(request, 's21_suplementacion6/index_s21_suplementacion6.html', {
        'red': red,
        'mes_seleccionado': mes_seleccionado,
    })

## SEGUIMIENTO
def get_redes_s21_suplementacion6(request,redes_id):
    redes = (
            MAESTRO_HIS_ESTABLECIMIENTO
            .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL',Departamento='JUNIN')
            .annotate(codigo_red_filtrado=Substr('Codigo_Red', 1, 4))
            .values('Red','codigo_red_filtrado')
            .distinct()
            .order_by('Red')
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'redes': redes,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
    }
    
    return render(request, 's21_suplementacion6/redes.html', context)

def obtener_seguimiento_redes_s21_suplementacion6(p_red,p_inicio,p_fin):
    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT DISTINCT * FROM public.fn_seguimiento_s21_suplementacion6(%s, %s, %s)",
            [p_red, p_inicio, p_fin]
        )
        return cursor.fetchall()

class RptS21Suplementacion6Red(TemplateView):
    def get(self, request, *args, **kwargs):
        # Variables ingresadas
        p_red = request.GET.get('red')
        p_inicio = request.GET.get('fecha_inicio')
        p_fin = request.GET.get('fecha_fin')
        # Creación de la consulta
        resultado_seguimiento = obtener_seguimiento_redes_s21_suplementacion6(p_red, p_inicio, p_fin)
        
        wb = Workbook()
        
        consultas = [
                ('Seguimiento', resultado_seguimiento)
        ]
        
        for index, (sheet_name, results) in enumerate(consultas):
            if index == 0:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
        
            fill_worksheet_s21_suplementacion6(ws, results)
        ##########################################################################          
        # Establecer el nombre del archivo
        nombre_archivo = "rpt_s21_suplementacion6.xlsx"
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)

        return response

def fill_worksheet_s21_suplementacion6(ws, results): 
    # cambia el alto de la columna
    ws.row_dimensions[1].height = 14
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 3
    ws.row_dimensions[4].height = 25
    ws.row_dimensions[5].height = 3
    ws.row_dimensions[7].height = 14
    ws.row_dimensions[8].height = 25
    # cambia el ancho de la columna
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 9
    ws.column_dimensions['C'].width = 9
    ws.column_dimensions['D'].width = 6
    ws.column_dimensions['E'].width = 9
    ws.column_dimensions['F'].width = 5
    ws.column_dimensions['G'].width = 6
    ws.column_dimensions['H'].width = 6
    ws.column_dimensions['I'].width = 9
    ws.column_dimensions['J'].width = 5
    ws.column_dimensions['K'].width = 10
    ws.column_dimensions['L'].width = 9
    ws.column_dimensions['M'].width = 5
    ws.column_dimensions['N'].width = 9
    ws.column_dimensions['O'].width = 5
    ws.column_dimensions['P'].width = 9
    ws.column_dimensions['Q'].width = 5
    ws.column_dimensions['R'].width = 9
    ws.column_dimensions['S'].width = 5
    ws.column_dimensions['T'].width = 9
    ws.column_dimensions['U'].width = 5
    ws.column_dimensions['V'].width = 9
    ws.column_dimensions['W'].width = 9
    ws.column_dimensions['X'].width = 5
    ws.column_dimensions['Y'].width = 9
    ws.column_dimensions['Z'].width = 5
    ws.column_dimensions['AA'].width = 10
    ws.column_dimensions['AB'].width = 11
    ws.column_dimensions['AC'].width = 11
    ws.column_dimensions['AD'].width = 9
    ws.column_dimensions['AE'].width = 16
    ws.column_dimensions['AF'].width = 16
    ws.column_dimensions['AG'].width = 20
    ws.column_dimensions['AH'].width = 20
    ws.column_dimensions['AI'].width = 6
    ws.column_dimensions['AJ'].width = 25
    
    # linea de division
    ws.freeze_panes = 'L9'
    
    # Configuración del fondo y el borde
    # Definir el color usando formato aRGB (opacidad completa 'FF' + color RGB)
    fill = PatternFill(start_color='FF60D7E0', end_color='FF60D7E0', fill_type='solid')
    # Definir el color anaranjado usando formato aRGB
    orange_fill = PatternFill(start_color='FFE0A960', end_color='FFE0A960', fill_type='solid')
    # Definir los estilos para gris
    gray_fill = PatternFill(start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid')
    # Definir el estilo de color verde
    green_fill = PatternFill(start_color='FF60E0B3', end_color='FF60E0B3', fill_type='solid')
    # Definir el estilo de color amarillo
    yellow_fill = PatternFill(start_color='FFE0DE60', end_color='FFE0DE60', fill_type='solid')
    # Definir el estilo de color azul
    blue_fill = PatternFill(start_color='FF60A2E0', end_color='FF60A2E0', fill_type='solid')
    # Definir el estilo de color verde 2
    green_fill_2 = PatternFill(start_color='FF60E07E', end_color='FF60E07E', fill_type='solid')   
    
    green_font = Font(name='Arial', size=8, color='00FF00')  # Verde
    red_font = Font(name='Arial', size=8, color='FF0000')    # Rojo
    
    border = Border(left=Side(style='thin', color='00B0F0'),
                    right=Side(style='thin', color='00B0F0'),
                    top=Side(style='thin', color='00B0F0'),
                    bottom=Side(style='thin', color='00B0F0'))
    borde_plomo = Border(left=Side(style='thin', color='A9A9A9'), # Plomo
                    right=Side(style='thin', color='A9A9A9'), # Plomo
                    top=Side(style='thin', color='A9A9A9'), # Plomo
                    bottom=Side(style='thin', color='A9A9A9')) # Plomo
    
    ## crea titulo del reporte
    ws['B1'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B1'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['B1'] = 'OFICINA DE TECNOLOGIAS DE LA INFORMACION'
    
    ws['B2'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B2'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['B2'] = 'DIRECCION REGIONAL DE SALUD JUNIN'
    
    ws['B4'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B4'].font = Font(name = 'Arial', size= 12, bold = True)
    ws['B4'] = 'SEGUIMIENTO NOMINAL DEL INDICADOR SI-02.01. PORCENTAJE DE NIÑAS Y NIÑOS MESES DE EDAD DEL DEPARTAMENTO, DE SEIS (06) MESES DE EDAD, PREMATUROS Y/O BAJO PESO AL NACER, Y DE CUATRO (04) SIN DIAGNÓSTICO DE ANEMIA; QUE RECIBEN DOSAJES DE HEMOGLOBINA Y CULMINAN LA SUPLEMENTACIÓN PREVENTIVA CON HIERRO'
    
    ws['B6'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B6'].font = Font(name = 'Arial', size= 7, bold = True, color='0000CC')
    ws['B6'] ='El usuario se compromete a mantener la confidencialidad de los datos personales que conozca como resultado del reporte realizado, cumpliendo con lo establecido en la Ley N° 29733 - Ley de Protección de Datos Personales y sus normas complementarias.'

    ws['L7'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['L7'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['L7'] = 'NIÑOS PREMATUROS Y/O BAJO PESO AL NACER'
    
    ws['W7'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['W7'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['W7'] = 'NIÑOS SIN PREMATURIDAD'
    
    ws['B8'].alignment = Alignment(horizontal= "center", vertical="center")
    ws['B8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['B8'].fill = fill
    ws['B8'].border = border
    ws['B8'] = 'NUM DOC'
        
    ws['C8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['C8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['C8'].fill = fill
    ws['C8'].border = border
    ws['C8'] = 'FECHA NAC'      
    
    ws['D8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['D8'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
    ws['D8'].fill = fill
    ws['D8'].border = border
    ws['D8'] = 'SEXO' 
    
    ws['E8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['E8'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
    ws['E8'].fill = fill
    ws['E8'].border = border
    ws['E8'] = 'SEGURO'     
    
    ws['F8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['F8'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
    ws['F8'].fill = fill
    ws['F8'].border = border
    ws['F8'] = 'CNV'    
    
    ws['G8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['G8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['G8'].fill = fill
    ws['G8'].border = border
    ws['G8'] = 'PESO'    
    
    ws['H8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['H8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['H8'].fill = fill
    ws['H8'].border = border
    ws['H8'] = 'SEM GEST'   
    
    ws['I8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['I8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['I8'].fill = fill
    ws['I8'].border = border
    ws['I8'] = 'DX ANEMIA'
    
    ws['J8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['J8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['J8'].fill = fill
    ws['J8'].border = border
    ws['J8'] = 'VAL DX'         
    
    ws['K8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['K8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['K8'].fill = blue_fill
    ws['K8'].border = border
    ws['K8'] = 'DEN IND'    
    
    ws['L8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['L8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['L8'].fill = green_fill
    ws['L8'].border = border
    ws['L8'] = 'HB 30D'    
    
    ws['M8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['M8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['M8'].fill = green_fill
    ws['M8'].border = border
    ws['M8'] = 'VAL'    
    
    ws['N8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['N8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['N8'].fill = green_fill
    ws['N8'].border = border
    ws['N8'] = 'SUP 30D'    
    
    ws['O8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['O8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['O8'].fill = green_fill
    ws['O8'].border = border
    ws['O8'] = 'VAL'    
    
    ws['P8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['P8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['P8'].fill = green_fill
    ws['P8'].border = border
    ws['P8'] = 'HB 3M'    
    
    ws['Q8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['Q8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['Q8'].fill = green_fill
    ws['Q8'].border = border
    ws['Q8'] = 'VAL'    
    
    ws['R8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['R8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['R8'].fill = green_fill
    ws['R8'].border = border
    ws['R8'] = 'SUP 4M'  
    
    ws['S8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['S8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['S8'].fill = green_fill
    ws['S8'].border = border
    ws['S8'] = 'VAL'  
    
    ws['T8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['T8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['T8'].fill = green_fill
    ws['T8'].border = border
    ws['T8'] = 'HB 6M'  
    
    ws['U8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['U8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['U8'].fill = green_fill
    ws['U8'].border = border
    ws['U8'] = 'VAL'  
    
    ws['V8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['V8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['V8'].fill = blue_fill
    ws['V8'].border = border
    ws['V8'] = 'IND BPN'  
    
    ws['W8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['W8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['W8'].fill = green_fill_2
    ws['W8'].border = border
    ws['W8'] = 'SUP 4M' 
    
    ws['X8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['X8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['X8'].fill = green_fill_2
    ws['X8'].border = border
    ws['X8'] = 'VAL'  
    
    ws['Y8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['Y8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['Y8'].fill = green_fill_2
    ws['Y8'].border = border
    ws['Y8'] = 'HB 6M' 
    
    ws['Z8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['Z8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['Z8'].fill = green_fill_2
    ws['Z8'].border = border
    ws['Z8'] = 'VAL'  
    
    ws['AA8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AA8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['AA8'].fill = blue_fill
    ws['AA8'].border = border
    ws['AA8'] = 'IND SIN BPN' 
            
    ws['AB8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AB8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['AB8'].fill = fill
    ws['AB8'].border = border
    ws['AB8'] = 'MES' 
    
    ws['AC8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AC8'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AC8'].fill = gray_fill
    ws['AC8'].border = border
    ws['AC8'] = 'IND' 
    
    ws['AD8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AD8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['AD8'].fill = orange_fill
    ws['AD8'].border = border
    ws['AD8'] = 'UBIGEO'  
    
    ws['AE8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AE8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['AE8'].fill = orange_fill
    ws['AE8'].border = border
    ws['AE8'] = 'PROVINCIA'       
    
    ws['AF8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AF8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['AF8'].fill = orange_fill
    ws['AF8'].border = border
    ws['AF8'] = 'DISTRITO' 
    
    ws['AG8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AG8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['AG8'].fill = orange_fill
    ws['AG8'].border = border
    ws['AG8'] = 'RED'  
    
    ws['AH8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AH8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['AH8'].fill = orange_fill
    ws['AH8'].border = border
    ws['AH8'] = 'MICRORED'  
    
    ws['AI8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AI8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['AI8'].fill = orange_fill
    ws['AI8'].border = border
    ws['AI8'] = 'COD EST'  
    
    ws['AJ8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AJ8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['AJ8'].fill = orange_fill
    ws['AJ8'].border = border
    ws['AJ8'] = 'ESTABLECIMIENTO'  
    
    # Definir estilos
    header_font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    centered_alignment = Alignment(horizontal='center')
    border = Border(left=Side(style='thin', color='A9A9A9'),
            right=Side(style='thin', color='A9A9A9'),
            top=Side(style='thin', color='A9A9A9'),
            bottom=Side(style='thin', color='A9A9A9'))
    header_fill = PatternFill(patternType='solid', fgColor='00B0F0')
    
    # Definir los caracteres especiales de check y X
    check_mark = '✓'  # Unicode para check
    x_mark = '✗'  # Unicode para X
    sub_cumple = 'CUMPLE'
    sub_no_cumple = 'NO CUMPLE'
    
    # Escribir datos
    for row, record in enumerate(results, start=9):
        for col, value in enumerate(record, start=2):
            cell = ws.cell(row=row, column=col, value=value)

            # Alinear a la izquierda solo en las columnas 6,14,15,16
            if col in [34, 36, 38]:
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='center')

            # Aplicar color en la columna 27
            if col == 29:
                if isinstance(value, str):
                    value_upper = value.strip().upper()
                    if value_upper == "NO CUMPLE":
                        cell.fill = PatternFill(patternType='solid', fgColor='FF0000')  # Fondo rojo
                        cell.font = Font(name='Arial', size=7,  bold = True, color='FFFFFF')  # Letra blanca
                    elif value_upper == "CUMPLE":
                        cell.fill = PatternFill(patternType='solid', fgColor='00FF00')  # Fondo verde
                        cell.font = Font(name='Arial', size=7,  bold = True, color='FFFFFF')  # Letra blanca
                    else:
                        cell.font = Font(name='Arial', size=7)
                else:
                    cell.font = Font(name='Arial', size=8)
            
            # Aplicar color de letra SUB INDICADORES
            elif col in [11, 22, 24, 27]:
                if value == 0:
                    cell.value = sub_no_cumple  # Insertar check
                    cell.font = Font(name='Arial', size=7, color="FF0000")  # Letra roja
                elif value == 1:
                    cell.value = sub_cumple # Insertar check
                    cell.font = Font(name='Arial', size=7, color="00B050")  # Letra verde
                else:
                        cell.font = Font(name='Arial', size=7)

            # Fuente normal para otras columnas
            else:
                cell.font = Font(name='Arial', size=8)  # Fuente normal para otras columnas

            # Aplicar caracteres especiales check y X
            if col in [6, 10, 13, 15, 17, 19, 21, 24, 26]:
                if value == 1:
                    cell.value = check_mark  # Insertar check
                    cell.font = Font(name='Arial', size=10, color='00B050')  # Letra verde
                elif value == 0:
                    cell.value = x_mark  # Insertar X
                    cell.font = Font(name='Arial', size=10, color='FF0000')  # Letra roja
                else:
                    cell.font = Font(name='Arial', size=8)  # Fuente normal si no es 1 o 0
            
                        
            cell.border = border